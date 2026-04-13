# ============================================================================
# IMPORTS
# ============================================================================
import os
import sys
import time
import re
from datetime import datetime, timedelta
import threading

# Ensure stdout is unbuffered for immediate output visibility in Docker
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

# Import the new logging system
from utils.logger import log_info, log_success, log_error, log_warning, log_debug, log_database
from flask import Flask, request, jsonify, render_template_string, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_cors import CORS
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from services.send_email import send_email
from services.doc_processing import process_doc_processing
from services.web_exploring import web_explorer_agent, get_company_data_with_priority_flow
import uuid
import json
from pathlib import Path
from config import config
from sqlalchemy.ext.mutable import MutableDict
from sqlalchemy import text

# ============================================================================
# THREAD MANAGEMENT
# ============================================================================
# Thread tracking system for background processes
class ThreadTracker:
    def __init__(self):
        self.active_threads = {}  # profile_id -> list of thread objects
        self.lock = threading.Lock()
    
    def add_thread(self, profile_id, thread):
        """Add a thread to the tracking system for a specific profile."""
        with self.lock:
            if profile_id not in self.active_threads:
                self.active_threads[profile_id] = []
            self.active_threads[profile_id].append(thread)
    
    def remove_thread(self, profile_id, thread):
        """Remove a thread from the tracking system."""
        with self.lock:
            if profile_id in self.active_threads:
                try:
                    self.active_threads[profile_id].remove(thread)
                    if not self.active_threads[profile_id]:
                        del self.active_threads[profile_id]
                        # print(f"🔗 No more active threads for profile {profile_id}", flush=True)
                except ValueError:
                    pass  # Thread was already removed
    
    def stop_profile_threads(self, profile_id):
        """Stop all threads associated with a specific profile."""
        with self.lock:
            if profile_id in self.active_threads:
                threads_to_stop = self.active_threads[profile_id].copy()
                print(f"🛑 Stopping {len(threads_to_stop)} threads for profile {profile_id}", flush=True)
                
                for thread in threads_to_stop:
                    if thread.is_alive():
                        # Set a flag to stop the thread gracefully
                        if hasattr(thread, '_stop_flag'):
                            thread._stop_flag = True
                        print(f"🛑 Stopping thread {thread.name} for profile {profile_id}", flush=True)
                
                # Clear the threads from tracking
                del self.active_threads[profile_id]
                print(f"🛑 Cleared thread tracking for profile {profile_id}", flush=True)
                
                return len(threads_to_stop)
            return 0
    
    def get_active_profiles(self):
        """Get list of profile IDs that have active threads."""
        with self.lock:
            return list(self.active_threads.keys())
    
    def cleanup_dead_threads(self):
        """Remove references to threads that are no longer alive."""
        with self.lock:
            for profile_id in list(self.active_threads.keys()):
                self.active_threads[profile_id] = [
                    thread for thread in self.active_threads[profile_id] 
                    if thread.is_alive()
                ]
                if not self.active_threads[profile_id]:
                    del self.active_threads[profile_id]

# Global thread tracker instance
thread_tracker = ThreadTracker()

# ============================================================================
# FLASK APP INITIALIZATION
# ============================================================================
app = Flask(__name__)

# Configuration
env = os.environ.get('FLASK_ENV', 'default')
app.config.from_object(config[env])

# Extensions
db = SQLAlchemy(app)
jwt = JWTManager(app)
CORS(app, origins=app.config['CORS_ORIGINS'])

# Ensure upload directory exists
Path(app.config['UPLOAD_FOLDER']).mkdir(exist_ok=True)

# ============================================================================
# BACKGROUND PROCESSING FUNCTIONS
# ============================================================================
def process_news_retrieval(flask_app, db, CompanyProfile, profile_id: str) -> None:
    """Process news retrieval for a company profile in a separate thread."""
    current_thread = threading.current_thread()
    
    print(f"🔍 DEBUG: process_news_retrieval called for profile {profile_id}", flush=True)
    
    with flask_app.app_context():
        try:
            # Register this thread with the tracker
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Check if we should stop before starting
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 News retrieval stopped for profile {profile_id} before starting", flush=True)
                return
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found for news retrieval", flush=True)
                return
            
            # Check if profile is ready for news processing (after KPIs are extracted)
            profile_data = profile.profile_data or {}
            processing_stage = profile_data.get('processing_stage', '')
            
            # Allow news processing once KPIs are extracted (kpis_extracted), confirmed (kpis_confirmed), generating_report, or completed
            if processing_stage not in ['kpis_extracted', 'kpis_confirmed', 'generating_report', 'completed']:
                print(f"📰 Profile {profile_id} not ready for news processing (stage: {processing_stage}), skipping", flush=True)
                return
            
            print(f"📰 Starting news retrieval for {profile.company_name} (stage: {processing_stage})", flush=True)
            
            # Call news retrieval
            from services.news_retrieving import news_retriever_agent
            import os
            anthropic_key = os.getenv("ANTHROPIC_API_KEY")
            news_data = news_retriever_agent(profile.company_name, anthropic_key)
            
            # Check stop flag after news retrieval
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 News retrieval stopped for profile {profile_id} after retrieval", flush=True)
                return
            
            # Refresh profile from database to avoid stale data issues
            try:
                db.session.refresh(profile)
                print(f"📰 Profile refreshed for news data update", flush=True)
            except Exception as refresh_error:
                print(f"⚠️ Warning: Could not refresh profile: {refresh_error}", flush=True)
            
            # Update profile with news data
            if profile.profile_data is None:
                profile.profile_data = {}
            
            # Store news data in profile_data for later use
            profile.profile_data['news_data'] = news_data
            
            try:
                db.session.commit()
                print(f"💾 News data committed successfully for profile {profile_id}", flush=True)
                
                # Log what we received
                if isinstance(news_data, dict):
                    analysis_len = len(news_data.get('analysis', ''))
                    urls_count = len(news_data.get('urls', []))
                    print(f"📰 News data saved: {analysis_len} chars analysis, {urls_count} URLs", flush=True)
                else:
                    print(f"📰 News data: {len(str(news_data))} chars", flush=True)
                    
            except Exception as commit_error:
                print(f"❌ ERROR: Failed to commit news data: {str(commit_error)}", flush=True)
                db.session.rollback()
                # Try to get fresh profile and retry once
                try:
                    fresh_profile = db.session.get(CompanyProfile, profile_id)
                    if fresh_profile:
                        if fresh_profile.profile_data is None:
                            fresh_profile.profile_data = {}
                        fresh_profile.profile_data['news_data'] = news_data
                        db.session.commit()
                        print(f"💾 News data committed on retry for profile {profile_id}", flush=True)
                    else:
                        print(f"❌ Profile {profile_id} no longer exists", flush=True)
                except Exception as retry_error:
                    print(f"❌ Retry also failed: {retry_error}", flush=True)
                    raise retry_error
                
        except Exception as e:
            print(f"❌ News retrieval failed for profile {profile_id}: {str(e)}", flush=True)
            # Try to save minimal news data structure
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile:
                    if profile.profile_data is None:
                        profile.profile_data = {}
                    profile.profile_data['news_data'] = {
                        'analysis': f'Actualités pour {profile.company_name} temporairement indisponibles.',
                        'urls': []
                    }
                    db.session.commit()
                    print(f"💾 Minimal news data saved for profile {profile_id}", flush=True)
            except Exception as save_error:
                print(f"❌ Failed to save minimal news data for profile {profile_id}: {str(save_error)}", flush=True)
        finally:
            # Always remove this thread from tracking when done
            thread_tracker.remove_thread(profile_id, current_thread)

def process_web_exploring_with_news_sync(flask_app, db, CompanyProfile, profile_id: str, news_thread) -> None:
    """Process web exploring for a company profile, waiting for news retrieval to complete first."""
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread with the tracker
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Check if we should stop before starting
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Web exploring stopped for profile {profile_id} before starting", flush=True)
                return
            
            # Wait for news retrieval to complete (with timeout)
            news_thread.join(timeout=300)  # 5 minutes timeout (increased due to login retries)
            
            # Check stop flag after waiting for news
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Web exploring stopped for profile {profile_id} after news wait", flush=True)
                return
            
            if news_thread.is_alive():
                print(f"⚠️ News retrieval timeout for profile {profile_id}, proceeding with web exploring", flush=True)
            else:
                print(f"✅ News retrieval completed, starting web exploring for profile {profile_id}", flush=True)
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found for web exploring", flush=True)
                return
            
            # Call web exploring with priority flow
            web_data = get_company_data_with_priority_flow(profile.company_name)
            
            # Check stop flag after web exploring
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                log_warning(f"Web exploring stopped for profile {profile_id} after web exploration")
                return
            
            # Debug: Log what web_data we received
            log_debug(f"Received web_data type: {type(web_data)}")
            if web_data:
                log_debug(f"Received web_data keys: {list(web_data.keys()) if isinstance(web_data, dict) else 'Not a dict'}")
            else:
                log_debug("Received web_data is None or empty!")
            
            # Update profile with web data
            if profile.profile_data is None:
                profile.profile_data = {}
            profile.profile_data['web_data'] = web_data
            
            try:
                db.session.commit()
                print(f"💾 Web data committed successfully to profile {profile_id}", flush=True)
                
                # Verify the data was actually saved
                print("profile: ", profile)
                
                # Try to refresh and verify, but handle potential issues gracefully
                try:
                    db.session.refresh(profile)
                    
                    # Add safety checks after refresh
                    if profile.profile_data is None:
                        print(f"⚠️  WARNING: profile_data is None after refresh!", flush=True)
                    elif 'web_data' not in profile.profile_data:
                        print(f"⚠️  WARNING: web_data key missing after refresh!", flush=True)
                        print(f"🔍 DEBUG: Available keys: {list(profile.profile_data.keys())}", flush=True)
                    else:
                        saved_web_data = profile.profile_data['web_data']
                        print(f"🔍 DEBUG: Verification - saved web data keys: {list(saved_web_data.keys())}", flush=True)
                        # print("saved_web_data: ", saved_web_data)
                        
                except Exception as refresh_error:
                    print(f"⚠️  WARNING: Error during refresh verification: {str(refresh_error)}", flush=True)
                    print(f"🔍 DEBUG: Web data was committed successfully, continuing despite refresh error", flush=True)
                
            except Exception as commit_error:
                print(f"❌ ERROR: Failed to commit web data: {str(commit_error)}", flush=True)
                db.session.rollback()
                raise commit_error
                
        except Exception as e:
            print(f"❌ Web exploring failed for profile {profile_id}: {str(e)}", flush=True)
            # Try to save minimal web data structure
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile:
                    if profile.profile_data is None:
                        profile.profile_data = {}
                    profile.profile_data['web_data'] = {
                        'basic_info': {
                            'companyOverview': {
                                'companyFoundationyear': 'Non spécifié',
                                'companyExpertise': 'À déterminer',
                                'primary_sector': 'Secteur général',
                                'legal_form': 'SARL',
                                'companyDefinition': f'Entreprise {profile.company_name} - informations à compléter',
                                'staff_count': 'À préciser'
                            },
                            'sectors': [],
                            'markets': [],
                            'keyPeople': [],
                            'contact': {
                                'phone': 'Non disponible',
                                'email': 'Non disponible',
                                'address': 'Adresse à préciser',
                                'website': 'Non disponible'
                            }
                        },
                        'news': f'Actualités pour {profile.company_name} temporairement indisponibles.',
                        'news_urls': []
                    }
                    db.session.commit()
                    print(f"💾 Minimal web data saved for profile {profile_id}", flush=True)
            except Exception as save_error:
                print(f"❌ Failed to save minimal web data for profile {profile_id}: {str(save_error)}", flush=True)
        finally:
            # Always remove this thread from tracking when done
            thread_tracker.remove_thread(profile_id, current_thread)

def process_web_exploring(flask_app, db, CompanyProfile, profile_id: str) -> None:
    """Process web exploring for a company profile in a separate thread."""
    with flask_app.app_context():
        try:
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found for web exploring", flush=True)
                return
            
            # Call web exploring with priority flow
            web_data = get_company_data_with_priority_flow(profile.company_name)
            
            # Debug: Log what web_data we received
            print(f"🔍 DEBUG: Received web_data type: {type(web_data)}", flush=True)
            if web_data:
                print(f"🔍 DEBUG: Received web_data keys: {list(web_data.keys()) if isinstance(web_data, dict) else 'Not a dict'}", flush=True)
            else:
                print(f"🔍 DEBUG: Received web_data is None or empty!", flush=True)
            
            # Update profile with web data
            # print("profile.profile_data: ", profile.profile_data)
            if profile.profile_data is None:
                profile.profile_data = {}
                # print("profile.profile_data if none: ", profile.profile_data)
            profile.profile_data['web_data'] = web_data
            # print("profile.profile_data['web_data']  ", profile.profile_data['web_data'])
            
            # Debug: Log before commit
            print(f"🔍 DEBUG: About to commit web_data to database", flush=True)
            print(f"🔍 DEBUG: Before commit - profile_data type: {type(profile.profile_data)}", flush=True)
            print(f"🔍 DEBUG: Before commit - profile_data keys: {list(profile.profile_data.keys()) if profile.profile_data else 'None'}", flush=True)
            
            try:
                db.session.commit()
                print(f"💾 Web data committed successfully to profile {profile_id}", flush=True)
                
                # Verify the data was actually saved
                print("profile: ", profile)
                
                # Try to refresh and verify, but handle potential issues gracefully
                try:
                    db.session.refresh(profile)
                    
                    # Add safety checks after refresh
                    if profile.profile_data is None:
                        print(f"⚠️  WARNING: profile_data is None after refresh!", flush=True)
                    elif 'web_data' not in profile.profile_data:
                        print(f"⚠️  WARNING: web_data key missing after refresh!", flush=True)
                        print(f"🔍 DEBUG: Available keys: {list(profile.profile_data.keys())}", flush=True)
                    else:
                        saved_web_data = profile.profile_data['web_data']
                        print(f"🔍 DEBUG: Verification - saved recommendation length: {len(saved_web_data.get('recommendation', ''))}", flush=True)
                        # print("saved_web_data: ", saved_web_data)
                        
                except Exception as refresh_error:
                    print(f"⚠️  WARNING: Error during refresh verification: {str(refresh_error)}", flush=True)
                    print(f"🔍 DEBUG: Web data was committed successfully, continuing despite refresh error", flush=True)
                
            except Exception as commit_error:
                print(f"❌ ERROR: Failed to commit web data: {str(commit_error)}", flush=True)
                db.session.rollback()
                raise commit_error
            
        except Exception as e:
            print(f"❌ Web exploring failed for profile {profile_id}: {str(e)}", flush=True)

def send_pdf_report_email(profile_id: str) -> bool:
    """Send PDF report via email to the user who requested it."""
    try:
        with app.app_context():
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found for email sending", flush=True)
                return False
                
            # Check if user opted for email report
            profile_data = profile.profile_data or {}
            email_report = profile_data.get('email_report', False)
            
            if not email_report:
                print(f"📧 User did not opt for email report for profile {profile_id}", flush=True)
                return False
                
            # Get user email
            user = db.session.get(User, profile.created_by)
            if not user or not user.email:
                print(f"❌ User email not found for profile {profile_id}", flush=True)
                return False
                
            # Generate PDF using the existing get_profile_pdf logic
            from datetime import datetime
            from io import BytesIO
            
            # Reuse the PDF generation logic from get_profile_pdf
            if profile.status != 'completed':
                print(f"⏳ Profile {profile_id} is not completed yet, skipping email", flush=True)
                return False
                
            extracted_kpis = profile_data.get('extracted_kpis')
            computed_ratios = profile_data.get('computed_ratios')
            web_data = profile_data.get('web_data', {})
            company_name = profile_data.get('company_name') or profile.company_name
            
            if not extracted_kpis and not computed_ratios:
                print(f"❌ Financial data not available for profile {profile_id}", flush=True)
                return False
                
            # Generate PDF content (simplified version)
            basic_info = web_data.get('basic_info', {})
            company_overview = basic_info.get('companyOverview', {})
            contact = basic_info.get('contact', {})
            
            def safe_get(value, default=''):
                return str(value) if value is not None else default
            
            # Create clean PDF-friendly HTML
            clean_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{company_name} - Company Profile</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 40px;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .header {{
                        text-align: center;
                        border-bottom: 3px solid #007bff;
                        padding-bottom: 20px;
                        margin-bottom: 30px;
                    }}
                    h1 {{
                        color: #007bff;
                        font-size: 28px;
                        margin-bottom: 10px;
                    }}
                    h2 {{
                        color: #555;
                        font-size: 20px;
                        margin-top: 30px;
                        margin-bottom: 15px;
                        border-left: 4px solid #007bff;
                        padding-left: 15px;
                    }}
                    .section {{
                        margin: 25px 0;
                        padding: 15px;
                        border: 1px solid #eee;
                        border-radius: 5px;
                    }}
                    .metric {{
                        margin: 8px 0;
                        padding: 5px 0;
                    }}
                    .metric strong {{
                        color: #007bff;
                        display: inline-block;
                        width: 150px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 15px 0;
                    }}
                    th, td {{
                        border: 1px solid #ddd;
                        padding: 8px;
                        text-align: left;
                    }}
                    th {{
                        background-color: #f2f2f2;
                        font-weight: bold;
                    }}
                    .footer {{
                        margin-top: 40px;
                        padding-top: 20px;
                        border-top: 1px solid #ddd;
                        font-size: 12px;
                        color: #666;
                        text-align: center;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{company_name}</h1>
                    <p>Company Profile Report</p>
                </div>

                <div class="section">
                    <h2>Company Overview</h2>
                    <div class="metric"><strong>Company Name:</strong> {company_name}</div>
                    <div class="metric"><strong>Legal Form:</strong> {safe_get(company_overview.get('legal_form'), 'Not specified')}</div>
                    <div class="metric"><strong>Founded:</strong> {safe_get(company_overview.get('companyFoundationyear'), 'Not specified')}</div>
                    <div class="metric"><strong>Primary Sector:</strong> {safe_get(company_overview.get('primary_sector'), 'General sector')}</div>
                    <div class="metric"><strong>Expertise:</strong> {safe_get(company_overview.get('companyExpertise'), 'To be determined')}</div>
                </div>"""
            
            # Add financial data if available
            if extracted_kpis:
                clean_html += """
                <div class="section">
                    <h2>Financial Information</h2>
                    <table>
                        <tr><th>Metric</th><th>Value</th></tr>"""
                
                for key, value in extracted_kpis.items():
                    if value is not None:
                        clean_html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
                
                clean_html += "</table></div>"
            
            if computed_ratios:
                clean_html += """
                    <div class="section">
                        <h2>Financial Ratios</h2>
                        <table>
                            <tr><th>Ratio</th><th>Value</th></tr>"""
                
                for key, value in computed_ratios.items():
                    if value is not None:
                        clean_html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
                
                clean_html += "</table></div>"
            
            clean_html += f"""
                <div class="section">
                    <h2>Contact Information</h2>
                    <div class="metric"><strong>Address:</strong> {safe_get(contact.get('address'), 'Not available')}</div>
                    <div class="metric"><strong>Phone:</strong> {safe_get(contact.get('phone'), 'Not available')}</div>
                    <div class="metric"><strong>Email:</strong> {safe_get(contact.get('email'), 'Not available')}</div>
                    <div class="metric"><strong>Website:</strong> {safe_get(contact.get('website'), 'Not available')}</div>
                </div>
                
                <div class="footer">
                    <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p>Company Profile Agent - Automated Analysis Report</p>
                </div>
            </body>
            </html>
            """
            
            # Generate PDF from HTML
            from weasyprint import HTML
            pdf_buffer = BytesIO()
            html_doc = HTML(string=clean_html, encoding='utf-8', base_url='')
            html_doc.write_pdf(pdf_buffer)
            
            # Get PDF content
            pdf_content = pdf_buffer.getvalue()
            pdf_buffer.close()
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"{company_name.replace(' ', '_')}_report_{timestamp}.pdf"
            
            # Send email with PDF attachment
            email_body = f"""Hello,

            Your company analysis report for {company_name} is now ready!

            Please find attached the comprehensive PDF report with:
            - Company overview and financial analysis
            - Key performance indicators and ratios
            - Market insights and recommendations

            The report has been automatically generated based on the documents you provided.

            Best regards,
            Company Profile Agent

            ---
            This is an automated message. The report was generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.
            """
            
            print(f"📧 Attempting to send PDF report to {user.email} for profile {profile_id}", flush=True)
            success = send_email(
                to_email=user.email,
                subject=f'Your Company Analysis Report - {company_name}',
                body=email_body,
                attachment_data=pdf_content,
                attachment_filename=filename
            )
            
            if success:
                print(f"✅ PDF report sent successfully to {user.email} for profile {profile_id}", flush=True)
                return True
            else:
                print(f"❌ Failed to send PDF report to {user.email} for profile {profile_id}", flush=True)
                return False
                
    except Exception as e:
        print(f"❌ Error sending PDF report for profile {profile_id}: {str(e)}", flush=True)
        return False

def process_doc_processing_with_web_sync(flask_app, db, CompanyProfile, LiasseDocument, profile_id: str, web_thread) -> None:
    """Process documents but wait for web exploring to complete first to avoid data conflicts."""
    with flask_app.app_context():
        try:
            
            # Wait for web exploring to complete (with timeout)
            web_thread.join(timeout=300)  # 5 minutes timeout
            
            if web_thread.is_alive():
                print(f"⚠️ Web exploring timeout for profile {profile_id}, proceeding with doc processing", flush=True)
            
            # Now proceed with document processing
            from services.doc_processing import process_doc_processing
            process_doc_processing(flask_app, db, CompanyProfile, LiasseDocument, profile_id)
            
            # After doc processing completes, set status to kpis_extracted
            # The user will need to confirm KPIs before news retrieval and financial analysis
            profile = db.session.get(CompanyProfile, profile_id)
            if profile:
                profile.status = 'processing'  # Keep as processing until user confirms KPIs
                db.session.commit()
                print(f"✅ Profile {profile_id} KPIs extracted, waiting for user confirmation", flush=True)
            
            # Check if profile was completed and send email if user opted in
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile and profile.status == 'completed':
                    email_report = bool((profile.profile_data or {}).get('email_report'))
                    if email_report:
                        print(f"📧 Profile {profile_id} completed with email preference, sending PDF report...", flush=True)
                        # Send PDF report via email in a separate thread to avoid blocking
                        import threading
                        email_thread = threading.Thread(
                            target=send_pdf_report_email,
                            args=(profile_id,),
                            daemon=True
                        )
                        email_thread.start()
                        print(f"🚀 Started email thread for profile {profile_id}", flush=True)
                    else:
                        print(f"📧 Profile {profile_id} completed but no email preference set", flush=True)
                else:
                    print(f"📧 Profile {profile_id} status: {profile.status if profile else 'not found'}", flush=True)
            except Exception as email_error:
                print(f"⚠️ Error checking email preference for profile {profile_id}: {str(email_error)}", flush=True)
            
        except Exception as e:
            print(f"❌ Doc processing with web sync failed for profile {profile_id}: {str(e)}", flush=True)

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
def _categorize_kpi(kpi_key):
    """Categorize KPIs for better organization in the UI."""
    # Crédit-bail: Redevances de crédit-bail, Redevances > 1 an, Redevances < 1 an, Prix d'achat résiduel (check first)
    if any(term in kpi_key.lower() for term in ['redevances', 'redevanes', 'prix', 'crédit']):
        return 'Crédit-bail'
    
    # Bilan Actif: Trésorerie-Actif, Titres Valeurs de placement, Compte d'associés (Actif), Actif circulant
    elif any(term in kpi_key.lower() for term in ['trésorerie-actif', 'titres valeurs de placement', 'compte d\'associés (actif)', 'actif circulant']):
        return 'Bilan Actif'
    
    # Bilan Passif: Trésorerie-passif, Compte d'associés (Passif), Capitaux propres, Passif circulant, Dettes de financement
    elif any(term in kpi_key.lower() for term in ['trésorerie-passif', 'compte d\'associés (passif)', 'capitaux propres', 'passif circulant', 'dettes de financement']):
        return 'Bilan Passif'
    
    # CPC: Chiffres d'affaires, Résultat d'exploitation, Dotations d'exploitation, Reprises d'exploitation, Résultat Net
    elif any(term in kpi_key.lower() for term in ['chiffre', 'résultat d\'exploitation', 'dotations d\'exploitation', 'reprises d\'exploitation', 'résultat net']):
        return 'CPC'
    
    else:
        return 'CPC'

def wait_for_db(max_retries=30, delay=1):
    """Wait for database to be ready with improved error handling."""
    database_url = app.config.get('SQLALCHEMY_DATABASE_URI', 'Not configured')
    hostname = 'unknown'
    
    # Extract hostname from DATABASE_URL for better error messages
    if database_url and database_url != 'Not configured':
        try:
            from urllib.parse import urlparse
            parsed = urlparse(database_url)
            hostname = parsed.hostname or 'unknown'
        except Exception:
            pass
    
    for attempt in range(max_retries):
        try:
            with app.app_context():
                db.engine.connect()
                print(f"✅ Database connection successful on attempt {attempt + 1}")
                return True
        except Exception as e:
            error_msg = str(e)
            print(f"❌ Database connection attempt {attempt + 1}/{max_retries} failed")
            print(f"   Hostname: {hostname}")
            print(f"   Error: {error_msg}")
            
            # Provide helpful guidance for common issues
            if attempt == max_retries - 1:
                print("\n" + "="*60)
                print("⚠️  DATABASE CONNECTION FAILED - Troubleshooting Tips:")
                print("="*60)
                if hostname == 'db':
                    print("• Hostname 'db' is only available inside Docker Compose network")
                    print("• If running locally, set DATABASE_URL to use 'localhost' instead:")
                    print("  export DATABASE_URL='postgresql://postgres:postgres@localhost:5432/company_profile_db'")
                    print("• If using Docker, ensure database container is running:")
                    print("  docker-compose up -d db")
                elif 'localhost' in hostname or '127.0.0.1' in hostname:
                    print("• Ensure PostgreSQL is running on your local machine")
                    print("• Check if PostgreSQL is listening on port 5432")
                    print("• Verify database credentials in DATABASE_URL")
                print("="*60 + "\n")
                raise e
            
            if attempt < max_retries - 1:
                time.sleep(delay)
    
    return False

# ============================================================================
# DATABASE MODELS
# ============================================================================
class User(db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    email = db.Column(db.String(255), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), default='analyst')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class CompanyProfile(db.Model):
    __tablename__ = 'company_profiles'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    company_name = db.Column(db.String(255), nullable=False)
    fiscal_years = db.Column(db.String(20), nullable=True)  # Fiscal years column (can be single year or range like "2022-2023")
    profile_data = db.Column(MutableDict.as_mutable(db.JSON))  # 👈 this
    status = db.Column(db.String(50), default='processing')
    created_by = db.Column(db.String(36), db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class LiasseDocument(db.Model):
    __tablename__ = 'liasse_documents'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    profile_id = db.Column(db.String(36), db.ForeignKey('company_profiles.id', ondelete='CASCADE'))
    document_type = db.Column(db.String(100))
    file_name = db.Column(db.String(255))
    file_path = db.Column(db.String(500))
    file_size = db.Column(db.Integer)
    upload_status = db.Column(db.String(50), default='uploaded')
    ocr_status = db.Column(db.String(50), default='pending')
    extracted_data = db.Column(db.JSON)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ============================================================================
# ROUTES
# ============================================================================

# ----------------------------------------------------------------------------
# Health & System Routes
# ----------------------------------------------------------------------------
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()})

# ----------------------------------------------------------------------------
# Authentication Routes
# ----------------------------------------------------------------------------
@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email and password are required'}), 400
        
        user = User.query.filter_by(email=email, is_active=True).first()
        
        if user and check_password_hash(user.password_hash, password):
            access_token = create_access_token(
                identity=user.id,
                additional_claims={
                    'email': user.email,
                    'role': user.role,
                    'name': f"{user.first_name} {user.last_name}"
                }
            )
            
            return jsonify({
                'access_token': access_token,
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'name': f"{user.first_name} {user.last_name}",
                    'role': user.role
                }
            })
        
        return jsonify({'error': 'Invalid credentials'}), 401
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        
        # Check if user already exists
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'User already exists'}), 400
        
        # Create new user
        user = User(
            email=data['email'],
            password_hash=generate_password_hash(data['password']),
            first_name=data['first_name'],
            last_name=data['last_name'],
            role=data.get('role', 'analyst')
        )
        
        db.session.add(user)
        db.session.commit()
        
        return jsonify({'message': 'User created successfully'}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# ----------------------------------------------------------------------------
# Dashboard Routes
# ----------------------------------------------------------------------------
@app.route('/api/dashboard/stats', methods=['GET'])
@jwt_required()
def get_dashboard_stats():
    """Get dashboard statistics including counts and percentage changes"""
    try:
        from sqlalchemy import func, extract
        from datetime import datetime, timedelta
        
        # Get current date info
        now = datetime.utcnow()
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        last_month_end = current_month_start - timedelta(seconds=1)
        
        # Total profiles count
        total_profiles = CompanyProfile.query.count()
        
        # This month's profiles
        this_month_profiles = CompanyProfile.query.filter(
            CompanyProfile.created_at >= current_month_start
        ).count()
        
        # Last month's profiles for comparison
        last_month_profiles = CompanyProfile.query.filter(
            CompanyProfile.created_at >= last_month_start,
            CompanyProfile.created_at <= last_month_end
        ).count()
        
        # Processing profiles
        processing_profiles = CompanyProfile.query.filter(
            CompanyProfile.status == 'processing'
        ).count()
        
        # Completed profiles
        completed_profiles = CompanyProfile.query.filter(
            CompanyProfile.status == 'completed'
        ).count()
        
        # Failed profiles (for success rate calculation)
        failed_profiles = CompanyProfile.query.filter(
            CompanyProfile.status == 'failed'
        ).count()
        
        # Debug: Let's see all status values
        all_profiles = CompanyProfile.query.all()
        status_counts = {}
        for profile in all_profiles:
            status = profile.status
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print(f"[DASHBOARD DEBUG] Status counts: {status_counts}", flush=True)
        print(f"[DASHBOARD DEBUG] Completed: {completed_profiles}, Failed: {failed_profiles}, Processing: {processing_profiles}", flush=True)
        
        # Calculate percentage changes
        total_change = 0
        if last_month_profiles > 0:
            total_change = round(((this_month_profiles - last_month_profiles) / last_month_profiles) * 100, 1)
        elif this_month_profiles > 0:
            total_change = 100.0
        
        # Calculate success rate
        total_processed = completed_profiles + failed_profiles
        success_rate = 0
        if total_processed > 0:
            success_rate = round((completed_profiles / total_processed) * 100, 1)
        elif completed_profiles > 0:
            # If we have completed profiles but no failed ones, success rate is 100%
            success_rate = 100.0
        
        print(f"[DASHBOARD DEBUG] Total processed: {total_processed}, Success rate: {success_rate}%", flush=True)
        
        # Calculate month-over-month change for total profiles
        # Get profiles from 2 months ago for comparison
        two_months_ago_start = (last_month_start - timedelta(days=1)).replace(day=1)
        two_months_ago_end = last_month_start - timedelta(seconds=1)
        
        two_months_ago_profiles = CompanyProfile.query.filter(
            CompanyProfile.created_at >= two_months_ago_start,
            CompanyProfile.created_at <= two_months_ago_end
        ).count()
        
        total_change_percent = 0
        if two_months_ago_profiles > 0:
            total_change_percent = round(((last_month_profiles - two_months_ago_profiles) / two_months_ago_profiles) * 100, 1)
        elif last_month_profiles > 0:
            total_change_percent = 100.0
        
        result = {
            'total_profiles': total_profiles,
            'this_month': this_month_profiles,
            'processing': processing_profiles,
            'completed': completed_profiles,
            'total_change_percent': total_change_percent,
            'month_change_percent': total_change,
            'success_rate': success_rate
        }
        
        print(f"[DASHBOARD DEBUG] Returning stats: {result}", flush=True)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/chart-data', methods=['GET'])
@jwt_required()
def get_dashboard_chart_data():
    """Get chart data for dashboard visualizations"""
    try:
        from sqlalchemy import func, extract
        from datetime import datetime, timedelta
        from calendar import monthrange
        
        now = datetime.utcnow()
        
        # Get last 6 months of data for activity chart
        monthly_data = []
        for i in range(5, -1, -1):
            # Calculate month start/end
            target_date = now - timedelta(days=30 * i)
            month_start = target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            days_in_month = monthrange(month_start.year, month_start.month)[1]
            month_end = month_start.replace(day=days_in_month, hour=23, minute=59, second=59)
            
            # Count profiles for this month
            count = CompanyProfile.query.filter(
                CompanyProfile.created_at >= month_start,
                CompanyProfile.created_at <= month_end
            ).count()
            
            # Count completed profiles for this month
            completed = CompanyProfile.query.filter(
                CompanyProfile.created_at >= month_start,
                CompanyProfile.created_at <= month_end,
                CompanyProfile.status == 'completed'
            ).count()
            
            monthly_data.append({
                'month': month_start.strftime('%b'),
                'year': month_start.year,
                'profiles': count,
                'completed': completed
            })
        
        # Get status distribution for pie chart
        status_distribution = []
        statuses = ['completed', 'processing', 'failed', 'pending']
        status_colors = {
            'completed': '#10b981',  # green
            'processing': '#f59e0b',  # amber
            'failed': '#ef4444',  # red
            'pending': '#6b7280'  # gray
        }
        
        for status in statuses:
            count = CompanyProfile.query.filter(
                CompanyProfile.status == status
            ).count()
            if count > 0:
                status_distribution.append({
                    'name': status.capitalize(),
                    'value': count,
                    'color': status_colors.get(status, '#6b7280')
                })
        
        # Get daily activity for the last 14 days (for sparkline)
        daily_activity = []
        for i in range(13, -1, -1):
            day = now - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59)
            
            count = CompanyProfile.query.filter(
                CompanyProfile.created_at >= day_start,
                CompanyProfile.created_at <= day_end
            ).count()
            
            daily_activity.append({
                'day': day_start.strftime('%d %b'),
                'count': count
            })
        
        # Calculate weekly comparison
        this_week_start = now - timedelta(days=now.weekday())
        this_week_start = this_week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        last_week_start = this_week_start - timedelta(days=7)
        last_week_end = this_week_start - timedelta(seconds=1)
        
        this_week_count = CompanyProfile.query.filter(
            CompanyProfile.created_at >= this_week_start
        ).count()
        
        last_week_count = CompanyProfile.query.filter(
            CompanyProfile.created_at >= last_week_start,
            CompanyProfile.created_at <= last_week_end
        ).count()
        
        weekly_change = 0
        if last_week_count > 0:
            weekly_change = round(((this_week_count - last_week_count) / last_week_count) * 100, 1)
        elif this_week_count > 0:
            weekly_change = 100.0
        
        return jsonify({
            'monthly_activity': monthly_data,
            'status_distribution': status_distribution,
            'daily_activity': daily_activity,
            'weekly_stats': {
                'this_week': this_week_count,
                'last_week': last_week_count,
                'change_percent': weekly_change
            }
        })
        
    except Exception as e:
        print(f"[CHART DATA ERROR] {str(e)}", flush=True)
        return jsonify({'error': str(e)}), 500


# ----------------------------------------------------------------------------
# Profile Management Routes (CRUD)
# ----------------------------------------------------------------------------
@app.route('/api/profiles', methods=['GET'])
@jwt_required()
def get_profiles():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '')
        
        query = CompanyProfile.query
        
        if search:
            query = query.filter(CompanyProfile.company_name.ilike(f'%{search}%'))
        
        # Sort profiles: processing status first, then by created_at DESC (recent to old)
        profiles = query.order_by(
            db.case(
                (CompanyProfile.status == 'processing', 0),
                else_=1
            ),
            CompanyProfile.created_at.desc()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Clean company names for display
        from services.profile_verification import _normalize_company_name
        
        return jsonify({
            'profiles': [{
                'id': p.id,
                'company_name': _normalize_company_name(p.company_name),
                'fiscal_years': p.fiscal_years,
                'status': p.status,
                'created_at': p.created_at.isoformat(),
                'profile_data': p.profile_data or {}
            } for p in profiles.items],
            'total': profiles.total,
            'pages': profiles.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ----------------------------------------------------------------------------
# Profile Operations Routes
# ----------------------------------------------------------------------------
@app.route('/api/profiles/verify', methods=['POST'])
@jwt_required()
def verify_profile():
    """
    Verify if a profile already exists by analyzing the first page of uploaded documents.
    This endpoint processes all uploaded files to extract company name and fiscal years,
    then checks if a matching profile already exists in the database.
    """
    try:
        from services.profile_verification import verify_profile_before_creation
        
        # Check if files were uploaded
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400
        
        files = request.files.getlist('files')
        if not files or len(files) == 0:
            return jsonify({'error': 'No files provided'}), 400
        
        # Get optional company name from form data
        company_name = request.form.get('company_name', '').strip()
        if not company_name:
            company_name = None
        
        # Validate all files
        valid_files = []
        for file in files:
            if file.filename == '':
                continue
            valid_files.append(file)
        
        if not valid_files:
            return jsonify({'error': 'No valid files provided'}), 400
        
        print(f"[VERIFICATION] Processing {len(valid_files)} files for verification", flush=True)
        
        # Save all files temporarily for analysis
        import tempfile
        import os
        
        temp_file_paths = []
        try:
            for i, file in enumerate(valid_files):
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                    file.save(temp_file.name)
                    temp_file_paths.append(temp_file.name)
            
            # Get Anthropic API key from config
            api_key = app.config.get('ANTHROPIC_API_KEY')
            if not api_key:
                return jsonify({'error': 'Anthropic API key not configured'}), 500
            
            # Perform verification with all files (with timeout handling)
            try:
                verification_result = verify_profile_before_creation(
                    temp_file_paths, 
                    api_key, 
                    db, 
                    CompanyProfile,
                    company_name
                )
            except Exception as verification_error:
                print(f"[VERIFICATION] Error during verification: {verification_error}", flush=True)
                return jsonify({
                    'success': False,
                    'error': 'Verification process failed. Please try again or contact support.',
                    'error_type': 'verification_failed',
                    'extracted_info': None,
                    'existing_profile': None
                }), 500
            
            # Log verification result summary
            if verification_result.get('existing_profile'):
                print(f"[VERIFICATION] Existing profile found: {verification_result['existing_profile']['id']}", flush=True)
            else:
                print(f"[VERIFICATION] No existing profile found", flush=True)
            
            return jsonify(verification_result)
            
        finally:
            # Clean up all temporary files
            for temp_path in temp_file_paths:
                try:
                    os.unlink(temp_path)
                    # print(f"[VERIFICATION] Cleaned up temporary file: {temp_path}", flush=True)
                except:
                    pass
        
    except Exception as e:
        print(f"[VERIFICATION] Error in verify_profile endpoint: {str(e)}", flush=True)
        return jsonify({'error': f'Verification failed: {str(e)}'}), 500

@app.route('/api/profiles/<profile_id>/debug', methods=['GET'])
@jwt_required()
def debug_profile_state(profile_id):
    """Debug endpoint to check profile state"""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
            
        profile_data = profile.profile_data or {}
        
        # Clean company name for display
        from services.profile_verification import _normalize_company_name
        
        return jsonify({
            'profile_id': profile_id,
            'company_name': _normalize_company_name(profile.company_name),
            'status': profile.status,
            'processing_stage': profile_data.get('processing_stage', 'NOT_SET'),
            'has_extracted_kpis': bool(profile_data.get('extracted_kpis')),
            'has_computed_ratios': bool(profile_data.get('computed_ratios')),
            'kpis_count': len(profile_data.get('extracted_kpis', {})),
            'ratios_count': len(profile_data.get('computed_ratios', {})),
            'profile_data_keys': list(profile_data.keys()),
            'last_run_finished_at': profile_data.get('last_run_finished_at')
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/kpis', methods=['GET'])
@jwt_required()
def get_profile_kpis(profile_id):
    """Get extracted KPIs for review by the user."""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
            
        profile_data = profile.profile_data or {}
        extracted_kpis = profile_data.get('extracted_kpis', {})
        processing_stage = profile_data.get('processing_stage', '')
        
        # Debug: Log the current state
        print(f"🔍 DEBUG: KPI endpoint called for profile {profile_id}", flush=True)
        print(f"🔍 DEBUG: Current stage: {processing_stage}", flush=True)
        print(f"🔍 DEBUG: Profile status: {profile.status}", flush=True)
        print(f"🔍 DEBUG: KPIs count: {len(extracted_kpis)}", flush=True)
        
        # Check if KPIs are available for review
        if processing_stage != 'kpis_extracted':
            # Provide more detailed error message based on current stage
            stage_messages = {
                'initializing': 'Document processing is initializing...',
                'waiting_for_documents': 'Waiting for documents to be uploaded...',
                'processing_documents': 'Documents are being processed...',
                'saving_markdown': 'Saving processed document content...',
                'extracting_kpis': 'Financial KPIs are being extracted...',
                'failed': 'KPI extraction failed. Please check the documents and try again.',
                '': 'KPI extraction has not started yet.'
            }
            
            # Handle dynamic stage names
            error_message = stage_messages.get(processing_stage, 'KPIs not yet extracted')
            if processing_stage and processing_stage.startswith('processing_document_'):
                match = re.match(r'processing_document_(\d+)_of_(\d+)', processing_stage)
                if match:
                    error_message = f'Processing document {match.group(1)} of {match.group(2)}...'
            
            return jsonify({
                'error': error_message,
                'processing_stage': processing_stage,
                'status': profile.status,
                'ready': False
            }), 400
            
        # Structure the KPIs for the frontend table
        kpi_structure = []
        kpi_mapping = {
            "Chiffre d'affaires": "Chiffre d'affaires",
            "Résultat d'exploitation": "Résultat d'exploitation", 
            "Résultat Net": "Résultat Net",
            "Dotations d'exploitation": "Dotations d'exploitation",
            "Reprises d'exploitation; transferts de charges": "Reprises d'exploitation",
            "Redevances de crédit-bail": "Redevances de crédit-bail",
            "Trésorerie-Actif": "Trésorerie-Actif",
            "Titres Valeurs de placement": "Titres Valeurs de placement",
            "Dettes de financement": "Dettes de financement",
            "Trésorerie-passif": "Trésorerie-passif",
            "Compte d'associés (Actif)": "Compte d'associés (Actif)",
            "Compte d'associés (Passif)": "Compte d'associés (Passif)",
            "Redevanes restant à payer (a plus d'un an)": "Redevances > 1 an",
            "Redevanes restant à payer (a moins d'un an)": "Redevances < 1 an",
            "Prix d'achat résiduel en fin du contrat": "Prix d'achat résiduel",
            "Capitaux propres": "Capitaux propres",
            "Actif circulant": "Actif circulant",
            "Passif circulant": "Passif circulant"
        }
        
        for kpi_key, display_name in kpi_mapping.items():
            kpi_data = extracted_kpis.get(kpi_key, {})
            
            # Handle both structured and unstructured data
            if isinstance(kpi_data, dict):
                n_value = kpi_data.get('N', kpi_data.get('n', None))
                n_minus_1_value = kpi_data.get('N-1', kpi_data.get('n-1', None))
            else:
                # Handle scalar values (fallback)
                n_value = kpi_data if kpi_data else None
                n_minus_1_value = None
                
            kpi_structure.append({
                'key': kpi_key,
                'display_name': display_name,
                'n_value': n_value,
                'n_minus_1_value': n_minus_1_value,
                'category': _categorize_kpi(kpi_key)
            })
        
        # Extract web data for sector and address validation
        web_data = profile_data.get('web_data', {})
        basic_info = web_data.get('basic_info', {})
        company_overview = basic_info.get('companyOverview', {})
        contact_info = basic_info.get('contact', {})
        
        # Extract sector and address information
        primary_sector = company_overview.get('primary_sector', 'Non spécifié')
        address = contact_info.get('address', 'Non spécifié')
        website = contact_info.get('website', 'Non spécifié')
        legal_form = company_overview.get('legal_form', 'Non spécifié')
        bizafrix_url = basic_info.get('bizafrix_url', '')
        
        # Clean company name for display
        from services.profile_verification import _normalize_company_name
        
        return jsonify({
            'success': True,
            'company_name': _normalize_company_name(profile.company_name),
            'fiscal_years': profile.fiscal_years,
            'kpis': kpi_structure,
            'processing_stage': processing_stage,
            'web_data': {
                'primary_sector': primary_sector,
                'address': address,
                'website': website,
                'legal_form': legal_form,
                'bizafrix_url': bizafrix_url
            }
        })
        
    except Exception as e:
        log_error(f"Error getting profile KPIs: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/kpis', methods=['PUT'])
@jwt_required()
def update_profile_kpis(profile_id):
    """Update KPIs based on user corrections."""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
            
        request_data = request.get_json()
        updated_kpis = request_data.get('kpis', [])
        updated_company_name = request_data.get('company_name', '').strip()
        updated_web_data = request_data.get('web_data', {})
        
        if not updated_kpis:
            return jsonify({'error': 'No KPIs provided'}), 400
            
        profile_data = profile.profile_data or {}
        extracted_kpis = profile_data.get('extracted_kpis', {})
        
        # Update the KPIs with user corrections
        for kpi_update in updated_kpis:
            kpi_key = kpi_update.get('key')
            n_value = kpi_update.get('n_value')
            n_minus_1_value = kpi_update.get('n_minus_1_value')
            
            if kpi_key:
                extracted_kpis[kpi_key] = {
                    'N': n_value,
                    'N-1': n_minus_1_value
                }
        
        # Update company name if provided
        if updated_company_name and updated_company_name != profile.company_name:
            profile.company_name = updated_company_name
            # Also update the company name in profile_data so report generation uses the updated name
            profile_data['company_name'] = updated_company_name
            log_info(f"Company name updated from '{profile.company_name}' to '{updated_company_name}' for profile {profile_id}")
        
        # Track if significant changes were made that require web exploring and news retrieval
        significant_changes_made = False
        bizafrix_re_scraping_triggered = False  # Track if Bizafrix re-scraping was triggered
        
        # Update web data if provided
        if updated_web_data:
            web_data = profile_data.get('web_data', {})
            basic_info = web_data.get('basic_info', {})
            company_overview = basic_info.get('companyOverview', {})
            contact_info = basic_info.get('contact', {})
            
            # Check for significant changes before updating
            old_primary_sector = company_overview.get('primary_sector', '')
            old_address = contact_info.get('address', '')
            old_legal_form = company_overview.get('legal_form', '')
            
            # Update the web data with user corrections
            if 'primary_sector' in updated_web_data:
                new_sector = updated_web_data['primary_sector']
                if new_sector != old_primary_sector:
                    significant_changes_made = True
                    log_info(f"Primary sector changed from '{old_primary_sector}' to '{new_sector}' for profile {profile_id}")
                company_overview['primary_sector'] = new_sector
                # Update company definition to reflect the new sector
                _update_company_definition(company_overview, new_sector, updated_company_name)
                
            if 'address' in updated_web_data:
                new_address = updated_web_data['address']
                if new_address != old_address:
                    significant_changes_made = True
                    log_info(f"Address changed from '{old_address}' to '{new_address}' for profile {profile_id}")
                contact_info['address'] = new_address
                
            if 'website' in updated_web_data:
                contact_info['website'] = updated_web_data['website']
                
            if 'legal_form' in updated_web_data:
                new_legal_form = updated_web_data['legal_form']
                if new_legal_form != old_legal_form:
                    significant_changes_made = True
                    log_info(f"Legal form changed from '{old_legal_form}' to '{new_legal_form}' for profile {profile_id}")
                company_overview['legal_form'] = new_legal_form
                
            if 'bizafrix_url' in updated_web_data:
                new_bizafrix_url = updated_web_data['bizafrix_url']
                old_bizafrix_url = basic_info.get('bizafrix_url', '')
                if new_bizafrix_url != old_bizafrix_url and new_bizafrix_url and new_bizafrix_url.strip():
                    # Bizafrix URL changed - trigger re-scraping directly instead of web exploring
                    log_info(f"Bizafrix URL changed from '{old_bizafrix_url}' to '{new_bizafrix_url}' for profile {profile_id}")
                    log_info(f"Triggering Bizafrix re-scraping for profile {profile_id}")
                    
                    # Store the URL first
                    basic_info['bizafrix_url'] = new_bizafrix_url
                    
                    # Trigger Bizafrix re-scraping in a separate thread
                    bizafrix_thread = threading.Thread(
                        target=process_bizafrix_re_scraping,
                        args=(app, db, CompanyProfile, profile_id, new_bizafrix_url),
                        daemon=False
                    )
                    bizafrix_thread.start()
                    log_info(f"Bizafrix re-scraping started for profile {profile_id}")
                    
                    # Mark that Bizafrix re-scraping was triggered (don't trigger web exploring)
                    bizafrix_re_scraping_triggered = True
                    # Don't mark significant_changes_made - re-scraping handles everything
                    # But we still need to update the profile data
                else:
                    # Just update the URL if it didn't change significantly
                    basic_info['bizafrix_url'] = new_bizafrix_url
            
            # Update the nested structure
            basic_info['companyOverview'] = company_overview
            basic_info['contact'] = contact_info
            web_data['basic_info'] = basic_info
            profile_data['web_data'] = web_data
            
            log_info(f"Web data updated for profile {profile_id}")
        
        # Check if company name changed (also significant)
        if updated_company_name and updated_company_name != profile.company_name:
            significant_changes_made = True
            log_info(f"Company name changed from '{profile.company_name}' to '{updated_company_name}' for profile {profile_id}")
        
        # Update the profile data
        profile_data['extracted_kpis'] = extracted_kpis
        profile_data['processing_stage'] = 'kpis_confirmed'
        
        # Recompute ratios with updated KPIs
        computed_ratios = _compute_financial_ratios(extracted_kpis)
        profile_data['computed_ratios'] = computed_ratios
        
        profile.profile_data = profile_data
        db.session.commit()
        
        # If significant changes were made, restart web exploring (but skip if Bizafrix re-scraping was triggered)
        if significant_changes_made and not bizafrix_re_scraping_triggered:
            log_info(f"Significant changes detected for profile {profile_id}, restarting web exploring")
            
            # Start web exploring with updated company information
            web_thread = threading.Thread(
                target=process_web_exploring_with_updated_data,
                args=(app, db, CompanyProfile, profile_id, updated_company_name or profile.company_name),
                daemon=False
            )
            web_thread.start()
            
            log_info(f"Web exploring restarted for profile {profile_id} due to significant changes")
            
            return jsonify({
                'success': True,
                'message': 'KPIs updated successfully. Web exploring restarted due to significant changes.',
                'processing_stage': 'kpis_confirmed',
                'web_exploring_restarted': True,
                'significant_changes': True
            })
        elif bizafrix_re_scraping_triggered:
            # Bizafrix re-scraping was triggered - return appropriate message
            log_info(f"Bizafrix re-scraping triggered for profile {profile_id}")
            return jsonify({
                'success': True,
                'message': 'KPIs updated successfully. Bizafrix re-scraping started with provided URL.',
                'processing_stage': 'kpis_confirmed',
                'bizafrix_re_scraping_triggered': True,
                'significant_changes': False
            })
        else:
            # No significant changes
            log_info(f"KPIs updated for profile {profile_id} by user {get_jwt_identity()}")
            
            return jsonify({
                'success': True,
                'message': 'KPIs updated successfully.',
                'processing_stage': 'kpis_confirmed',
                'significant_changes': False
            })
        
    except Exception as e:
        log_error(f"Error updating profile KPIs: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/generate-report', methods=['POST'])
@jwt_required()
def generate_full_report(profile_id):
    """Generate full report after KPI confirmation - starts news, web exploration, and analysis."""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
            
        profile_data = profile.profile_data or {}
        processing_stage = profile_data.get('processing_stage', '')
        
        # Allow report generation if user is on KPI review page (implicit validation)
        # Only block if processing has failed or is still in early stages
        if processing_stage in ['failed', 'processing_documents', 'extracting_kpis']:
            return jsonify({
                'error': 'Profile processing must be completed before generating report',
                'processing_stage': processing_stage
            }), 400
        
        # Check if Bizafrix URL exists but company name is still placeholder - trigger re-scraping first
        web_data = profile_data.get('web_data', {})
        basic_info = web_data.get('basic_info', {})
        bizafrix_url = basic_info.get('bizafrix_url', '')
        placeholder_values = ['company name placeholder', 'COMPANY NAME PLACEHOLDER', 
                             'Company Name Not Extracted', 'COMPANY NAME NOT EXTRACTED']
        is_placeholder = profile.company_name and profile.company_name.strip() in placeholder_values
        
        if bizafrix_url and bizafrix_url.strip() and is_placeholder:
            # Bizafrix URL provided but company name is still placeholder - trigger re-scraping
            print(f"🔗 Bizafrix URL found but company name is placeholder - triggering re-scraping before report generation", flush=True)
            log_info(f"Triggering Bizafrix re-scraping for profile {profile_id} before report generation")
            
            # Trigger Bizafrix re-scraping synchronously (wait for it to complete)
            # This ensures the company name is extracted before report generation
            try:
                process_bizafrix_re_scraping(app, db, CompanyProfile, profile_id, bizafrix_url)
                print(f"✅ Bizafrix re-scraping completed for profile {profile_id}", flush=True)
                
                # Refresh profile to get updated company name
                db.session.refresh(profile)
                profile_data = profile.profile_data or {}
                web_data = profile_data.get('web_data', {})
                basic_info = web_data.get('basic_info', {})
            except Exception as e:
                print(f"❌ Error in Bizafrix re-scraping before report generation: {e}", flush=True)
                log_error(f"Bizafrix re-scraping failed for profile {profile_id}: {e}")
                # Continue with report generation anyway
        
        # Check if news processing has already been done FIRST
        # News are considered processed if they exist and have content
        news_data = profile_data.get('news_data')
        news_already_processed = news_data is not None and (
            (isinstance(news_data, dict) and len(news_data) > 0) or 
            (isinstance(news_data, str) and len(news_data.strip()) > 0)
        )
        
        print(f"🔍 DEBUG: News already processed: {news_already_processed}", flush=True)
        print(f"🔍 DEBUG: Profile data keys: {list(profile_data.keys()) if profile_data else 'None'}", flush=True)
        if 'news_data' in profile_data:
            print(f"🔍 DEBUG: News data type: {type(profile_data['news_data'])}", flush=True)
        
        # Update status to generating_report immediately (before starting threads)
        profile_data['processing_stage'] = 'generating_report'
        profile.profile_data = profile_data
        profile.status = 'processing'
        db.session.commit()
        print(f"📊 Updated profile {profile_id} status to generating_report immediately", flush=True)
        
        if not news_already_processed:
            # Start news processing first (keep status as 'processing' for news retrieval)
            print(f"📰 Starting news processing for profile {profile_id} before report generation", flush=True)
            news_thread = threading.Thread(
                target=process_news_retrieval,
                args=(app, db, CompanyProfile, profile_id),
                daemon=False
            )
            news_thread.start()
            
            # Start final processing (will wait for news to complete)
            final_thread = threading.Thread(
                target=process_final_analysis_with_news_sync,
                args=(app, db, CompanyProfile, profile.id, news_thread),
                daemon=False
            )
            final_thread.start()
            
            log_info(f"News processing and final analysis started for profile {profile_id}")
        else:
            # News already processed, start final analysis directly (status already updated above)
            log_info(f"News processing already completed for profile {profile_id}, starting final analysis")
            final_thread = threading.Thread(
                target=process_final_analysis,
                args=(app, db, CompanyProfile, profile.id, None),  # No web_thread since web exploring is already done
                daemon=False
            )
            final_thread.start()
        
        log_info(f"Full report generation started for profile {profile_id}")
        
        return jsonify({
            'success': True,
            'processing_stage': 'generating_report'
        })
        
    except Exception as e:
        log_error(f"Error starting full report generation: {e}")
        return jsonify({'error': str(e)}), 500

# ----------------------------------------------------------------------------
# External Services Routes (Bizafrix, Charika)
# ----------------------------------------------------------------------------
@app.route('/api/extract-company-name-from-bizafrix', methods=['POST'])
@jwt_required()
def extract_company_name_from_bizafrix():
    """Extract company name from a Bizafrix URL."""
    try:
        data = request.get_json()
        bizafrix_url = data.get('bizafrix_url')
        
        if not bizafrix_url:
            return jsonify({'error': 'Bizafrix URL is required'}), 400
        
        # Validate URL format
        if not bizafrix_url.startswith('https://bizafrix.com/'):
            return jsonify({'error': 'Invalid Bizafrix URL format'}), 400
        
        # Extract company name from Bizafrix URL - DIRECT SCRAPING, NO SEARCH
        from services.bizafrix_web import get_bizafrix_company_details
        
        print(f"🔗 DIRECT SCRAPING: Extracting company name directly from Bizafrix URL (no search/scoring): {bizafrix_url}", flush=True)
        bizafrix_details = get_bizafrix_company_details(bizafrix_url)
        
        if not bizafrix_details:
            return jsonify({'error': 'Could not extract data from Bizafrix URL'}), 500
        
        company_name = bizafrix_details.get('company_name')
        
        if not company_name:
            return jsonify({'error': 'Company name not found in Bizafrix page'}), 404
        
        print(f"✅ Extracted company name: {company_name}", flush=True)
        
        return jsonify({
            'success': True,
            'company_name': company_name
        })
        
    except Exception as e:
        print(f"❌ Error extracting company name from Bizafrix: {e}", flush=True)
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/re-scrape-bizafrix', methods=['POST'])
@jwt_required()
def re_scrape_bizafrix(profile_id):
    """Re-scrape company information using a custom Bizafrix URL."""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
        
        data = request.get_json()
        bizafrix_url = data.get('bizafrix_url')
        
        if not bizafrix_url:
            return jsonify({'error': 'Bizafrix URL is required'}), 400
        
        # Validate URL format
        if not bizafrix_url.startswith('https://bizafrix.com/'):
            return jsonify({'error': 'Invalid Bizafrix URL format'}), 400
        
        # Start re-scraping in background
        re_scrape_thread = threading.Thread(
            target=process_bizafrix_re_scraping,
            args=(app, db, CompanyProfile, profile_id, bizafrix_url),
            daemon=True
        )
        re_scrape_thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Bizafrix re-scraping started successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/re-scrape-charika', methods=['POST'])
@jwt_required()
def re_scrape_charika(profile_id):
    """Re-scrape company information using a custom Charika URL."""
    try:
        profile = db.session.get(CompanyProfile, profile_id)
        if not profile:
            return jsonify({'error': 'Profile not found'}), 404
        
        data = request.get_json()
        charika_url = data.get('charika_url')
        
        if not charika_url:
            return jsonify({'error': 'Charika URL is required'}), 400
        
        # Validate URL format
        if not (charika_url.startswith('https://www.charika.ma/') or charika_url.startswith('https://charika.ma/')):
            return jsonify({'error': 'Invalid Charika URL format'}), 400
        
        # Start re-scraping in background
        re_scrape_thread = threading.Thread(
            target=process_charika_re_scraping,
            args=(app, db, CompanyProfile, profile_id, charika_url),
            daemon=True
        )
        re_scrape_thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Charika re-scraping started successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_final_analysis_with_news_sync(flask_app, db, CompanyProfile, profile_id: str, news_thread) -> None:
    """Process final financial analysis after news retrieval is complete."""
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Wait for news retrieval to complete (max 10 minutes)
            wait_timeout = 600
            if news_thread and news_thread.is_alive():
                print(f"⏳ Waiting for news retrieval to complete for profile {profile_id}...", flush=True)
                news_thread.join(timeout=wait_timeout)
            
            if news_thread and news_thread.is_alive():
                print(f"⚠️ News retrieval timeout for profile {profile_id}", flush=True)
            
            # Check stop flag
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Final analysis stopped for profile {profile_id}", flush=True)
                return
            
            # Status already updated to generating_report when endpoint was called
            # No need to update again here
            profile = CompanyProfile.query.get(profile_id)
            if profile:
                # Just refresh to ensure we have latest data
                db.session.refresh(profile)
                print(f"📊 Proceeding with final analysis for profile {profile_id} (status already set to generating_report)", flush=True)
            
            # Now proceed with final analysis
            process_final_analysis(flask_app, db, CompanyProfile, profile_id, None)
            
        except Exception as e:
            print(f"❌ Error in final analysis with news sync: {e}", flush=True)
        finally:
            thread_tracker.remove_thread(profile_id, current_thread)

def process_final_analysis(flask_app, db, CompanyProfile, profile_id: str, web_thread) -> None:
    """Process final financial analysis after all data is gathered."""
    import copy
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Wait for web exploring to complete (max 5 minutes)
            wait_timeout = 300
            if web_thread and web_thread.is_alive():
                print(f"⏳ Waiting for web exploring to complete for profile {profile_id}...", flush=True)
                web_thread.join(timeout=wait_timeout)
            
            if web_thread and web_thread.is_alive():
                print(f"⚠️ Web exploring timeout for profile {profile_id}", flush=True)
            
            # Check stop flag
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Final analysis stopped for profile {profile_id}", flush=True)
                return
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found", flush=True)
                return
            
            profile_data = profile.profile_data or {}
            extracted_kpis = profile_data.get('extracted_kpis', {})
            computed_ratios = profile_data.get('computed_ratios', {})
            news_data = profile_data.get('news_data', {})
            web_data = profile_data.get('web_data', {})
            
            # Debug: Log the web data being used for analysis
            print(f"📊 Web data for analysis: {web_data}", flush=True)
            if web_data and web_data.get('basic_info'):
                basic_info = web_data.get('basic_info', {})
                company_overview = basic_info.get('companyOverview', {})
                contact_info = basic_info.get('contact', {})
                print(f"📊 Analysis using - Primary Sector: {company_overview.get('primary_sector')}", flush=True)
                print(f"📊 Analysis using - Address: {contact_info.get('address')}", flush=True)
                print(f"📊 Analysis using - Website: {contact_info.get('website')}", flush=True)
                print(f"📊 Analysis using - Legal Form: {company_overview.get('legal_form')}", flush=True)
            
            # Generate financial analysis
            if extracted_kpis and computed_ratios:
                print(f"📊 Starting financial analysis for {profile.company_name}", flush=True)
                
                from services.financial_reporting import generate_financial_analysis
                
                financial_analysis = generate_financial_analysis(
                    company_name=profile.company_name,
                    extracted_kpis=extracted_kpis,
                    computed_ratios=computed_ratios,
                    news_data=news_data,
                    web_data=web_data,
                    fiscal_year=profile.fiscal_years
                )
                
                # Add financial analysis to profile data
                updated_profile_data = copy.deepcopy(profile_data)
                updated_profile_data.update(financial_analysis)
                updated_profile_data['processing_stage'] = 'completed'
                profile.status = 'completed'
                profile.profile_data = updated_profile_data
                db.session.commit()
                
                print(f"✅ Full analysis completed for {profile.company_name}", flush=True)
            else:
                print(f"❌ Missing required data for analysis: KPIs={bool(extracted_kpis)}, Ratios={bool(computed_ratios)}", flush=True)
                updated_profile_data = copy.deepcopy(profile_data)
                updated_profile_data['processing_stage'] = 'error'
                profile.status = 'failed'
                profile.profile_data = updated_profile_data
                db.session.commit()
            
        except Exception as e:
            print(f"❌ Error in final analysis: {e}", flush=True)
            import traceback
            print(traceback.format_exc(), flush=True)
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile:
                    existing_profile_data = copy.deepcopy(profile.profile_data) if profile.profile_data else {}
                    existing_profile_data['processing_stage'] = 'error'
                    profile.status = 'failed'
                    profile.profile_data = existing_profile_data
                    db.session.commit()
            except:
                pass
        finally:
            thread_tracker.remove_thread(profile_id, current_thread)

# Import the financial ratio computation function
from services.doc_processing import _compute_financial_ratios

def process_web_exploring_with_updated_data(flask_app, db, CompanyProfile, profile_id: str, company_name: str) -> None:
    """Process web exploring with updated company information."""
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Check stop flag
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Web exploring stopped for profile {profile_id}", flush=True)
                return
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found", flush=True)
                return
            
            print(f"🌐 Starting web exploring with updated data for {company_name}", flush=True)
            
            # Get updated web data from profile
            profile_data = profile.profile_data or {}
            web_data = profile_data.get('web_data', {})
            basic_info = web_data.get('basic_info', {})
            company_overview = basic_info.get('companyOverview', {})
            contact_info = basic_info.get('contact', {})
            
            # Use updated information for web exploring
            updated_primary_sector = company_overview.get('primary_sector', '')
            updated_address = contact_info.get('address', '')
            
            print(f"🌐 Web exploring with - Company: {company_name}, Sector: {updated_primary_sector}, Address: {updated_address}", flush=True)
            
            # Perform web exploring with updated company information and user hints
            new_web_data = get_company_data_with_priority_flow(company_name, updated_primary_sector, updated_address)
            
            if new_web_data:
                # Merge new web data with user corrections
                if 'basic_info' in new_web_data:
                    new_basic_info = new_web_data['basic_info']
                    new_company_overview = new_basic_info.get('companyOverview', {})
                    new_contact_info = new_basic_info.get('contact', {})
                    
                    # Preserve user corrections
                    if updated_primary_sector:
                        new_company_overview['primary_sector'] = updated_primary_sector
                        # Also preserve the user-updated company definition
                        old_company_definition = company_overview.get('companyDefinition', '')
                        if old_company_definition:
                            new_company_overview['companyDefinition'] = old_company_definition
                            print(f"📝 Preserved user-updated company definition: {old_company_definition[:100]}...", flush=True)
                    if updated_address:
                        new_contact_info['address'] = updated_address
                    
                    # Update the structure
                    new_basic_info['companyOverview'] = new_company_overview
                    new_basic_info['contact'] = new_contact_info
                    new_web_data['basic_info'] = new_basic_info
                
                # Update profile data with new web data
                profile_data['web_data'] = new_web_data
                profile.profile_data = profile_data
                db.session.commit()
                
                print(f"✅ Web exploring completed with updated data for {company_name}", flush=True)
            else:
                print(f"⚠️ Web exploring failed for {company_name}, keeping existing data", flush=True)
            
        except Exception as e:
            print(f"❌ Error in web exploring with updated data: {e}", flush=True)
        finally:
            thread_tracker.remove_thread(profile_id, current_thread)


def _update_company_definition(company_overview, new_sector, company_name):
    """
    Update the company definition to reflect the new primary sector.
    
    Args:
        company_overview (dict): The company overview data
        new_sector (str): The new primary sector
        company_name (str): The company name
    """
    try:
        # Generate a new company definition based on the updated sector
        # Use a more natural and professional description format
        company_definition = f"{company_name} est un acteur opérant dans le secteur de {new_sector}, avec un positionnement spécialisé sur le marché des entreprises et institutions marocaines. Son modèle repose sur une expertise sectorielle adaptée aux spécificités du secteur {new_sector}. L'entreprise se différencie par son savoir-faire dans le domaine {new_sector} et son ancrage local au Maroc. L'entreprise opère principalement dans le secteur d'activité {new_sector}."
        
        company_overview['companyDefinition'] = company_definition
        log_info(f"Company definition updated to reflect new sector: {new_sector}")
        
    except Exception as e:
        log_error(f"Error updating company definition: {e}")
        # Fallback to a simple definition
        company_overview['companyDefinition'] = f"Entreprise {company_name} spécialisée dans le secteur {new_sector}"

def process_bizafrix_re_scraping(flask_app, db, CompanyProfile, profile_id: str, bizafrix_url: str) -> None:
    """Re-scrape company information using a custom Bizafrix URL."""
    import copy
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Check stop flag
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Bizafrix re-scraping stopped for profile {profile_id}", flush=True)
                return
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found", flush=True)
                return
            
            # Stop any existing background threads for this profile to avoid concurrent scrapes
            try:
                stopped_count = thread_tracker.stop_profile_threads(profile_id)
                print(f"🛑 Stopped {stopped_count} existing threads before re-scraping Bizafrix for {profile_id}", flush=True)
            except Exception as stop_err:
                print(f"⚠️ Could not stop existing threads: {stop_err}", flush=True)
            
            print(f"🔗 Starting Bizafrix re-scraping with custom URL: {bizafrix_url}", flush=True)
            print(f"🔗 Profile company name: {profile.company_name}", flush=True)
            
            # Update processing stage
            import copy
            profile_data = copy.deepcopy(profile.profile_data) if profile.profile_data else {}
            profile_data['processing_stage'] = 're_scraping_bizafrix'
            profile.status = 'processing'
            profile.profile_data = profile_data
            db.session.commit()
            
            # Get company data using the custom Bizafrix URL directly
            from services.bizafrix_web import get_company_info_from_bizafrix, get_bizafrix_company_details, bizafrix_to_web_explorer_format
            
            # Use the custom URL for Bizafrix re-scraping - call DIRECTLY without search
            print(f"🔗 RE-SCRAPING: Calling get_bizafrix_company_details directly with URL: {bizafrix_url}", flush=True)
            
            # Get detailed information directly from the URL - NO SEARCH, NO COMPANY NAME MATCHING
            # When a direct URL is provided, we skip all search/matching and extract everything from the page
            print(f"🔗 EXECUTING DIRECT SCRAPING - NO SEARCH OR COMPANY NAME MATCHING", flush=True)
            print(f"🔗 Using direct URL provided by user: {bizafrix_url}", flush=True)
            bizafrix_details = get_bizafrix_company_details(bizafrix_url)
            print(f"🔗 Direct scraping completed, got {len(bizafrix_details) if bizafrix_details else 0} fields", flush=True)
            
            if not bizafrix_details:
                print(f"❌ Could not extract detailed data from Bizafrix for: {bizafrix_url}", flush=True)
                raise Exception("Could not extract data from provided URL")
            
            # Extract company name from the page - this is the authoritative source when URL is provided
            extracted_company_name = bizafrix_details.get("company_name")
            if not extracted_company_name:
                # Fallback: try to extract from URL slug if page extraction failed
                print(f"⚠️ Company name not found in page, trying URL slug fallback...", flush=True)
                try:
                    url_parts = bizafrix_url.rstrip('/').split('/')
                    if 'company' in url_parts:
                        company_index = url_parts.index('company')
                        if company_index + 1 < len(url_parts):
                            url_slug = url_parts[company_index + 1]
                            # Convert slug to readable name (replace hyphens with spaces and title case)
                            extracted_company_name = url_slug.replace('-', ' ').title()
                            print(f"🔗 Extracted company name from URL slug: {extracted_company_name}", flush=True)
                except Exception as url_parse_err:
                    print(f"⚠️ Could not extract company name from URL slug: {url_parse_err}", flush=True)
            
            if not extracted_company_name:
                raise Exception("Could not extract company name from Bizafrix URL page")
            
            # Always use the extracted company name from the page when URL is provided
            company_display_name = extracted_company_name
            print(f"📝 Using company name extracted from Bizafrix page: '{company_display_name}'", flush=True)
            
            # Always update the profile company_name with the extracted name (regardless of placeholder)
            # This ensures we replace any placeholder like "COMPANY NAME NOT EXTRACTED" with the real name
            if profile.company_name != company_display_name:
                print(f"📝 Updating profile company name from '{profile.company_name}' to '{company_display_name}'", flush=True)
                profile.company_name = company_display_name
                db.session.commit()
            
            # Convert to web explorer format
            print(f"🔗 Converting to web explorer format with company name: {company_display_name}", flush=True)
            bizafrix_data = bizafrix_to_web_explorer_format(bizafrix_details, company_display_name, bizafrix_url)
            
            # Add the Bizafrix URL to the result
            bizafrix_data['bizafrix_url'] = bizafrix_url
            
            print(f"✅ RE-SCRAPING COMPLETE: Successfully extracted Bizafrix data for '{company_display_name}' from URL: {bizafrix_url}", flush=True)
            
            # Wrap the data in the web_data format expected by the system
            new_web_data = {'basic_info': bizafrix_data} if bizafrix_data else None
            
            if new_web_data and new_web_data.get('basic_info'):
                # Update the profile with new web data
                updated_profile_data = copy.deepcopy(profile_data)
                updated_profile_data['web_data'] = new_web_data
                updated_profile_data['processing_stage'] = 'kpis_extracted'  # Ready for report generation
                profile.status = 'completed'
                profile.profile_data = updated_profile_data
                db.session.commit()
                
                print(f"✅ Bizafrix re-scraping completed successfully for {profile.company_name}", flush=True)
                
                # Automatically trigger report generation after successful re-scraping
                print(f"🚀 Auto-triggering report generation for {profile.company_name}", flush=True)
                report_thread = threading.Thread(
                    target=process_final_analysis,
                    args=(app, db, CompanyProfile, profile_id, None),
                    daemon=True
                )
                report_thread.start()
                
            else:
                print(f"❌ Bizafrix re-scraping failed for {profile.company_name}", flush=True)
                updated_profile_data = copy.deepcopy(profile_data)
                updated_profile_data['processing_stage'] = 'error'
                profile.status = 'failed'
                profile.profile_data = updated_profile_data
                db.session.commit()
            
        except Exception as e:
            print(f"❌ Error in Bizafrix re-scraping: {e}", flush=True)
            import traceback
            print(traceback.format_exc(), flush=True)
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile:
                    existing_profile_data = copy.deepcopy(profile.profile_data) if profile.profile_data else {}
                    existing_profile_data['processing_stage'] = 'error'
                    profile.status = 'failed'
                    profile.profile_data = existing_profile_data
                    db.session.commit()
            except:
                pass
        finally:
            thread_tracker.remove_thread(profile_id, current_thread)

def process_charika_re_scraping(flask_app, db, CompanyProfile, profile_id: str, charika_url: str) -> None:
    """Re-scrape company information using a custom Charika URL."""
    current_thread = threading.current_thread()
    
    with flask_app.app_context():
        try:
            # Register this thread
            thread_tracker.add_thread(profile_id, current_thread)
            
            # Check stop flag
            if hasattr(current_thread, '_stop_flag') and current_thread._stop_flag:
                print(f"🛑 Charika re-scraping stopped for profile {profile_id}", flush=True)
                return
            
            profile = db.session.get(CompanyProfile, profile_id)
            if not profile:
                print(f"❌ Profile {profile_id} not found", flush=True)
                return
            
            print(f"🔗 Starting Charika re-scraping with custom URL: {charika_url}", flush=True)
            print(f"🔗 Profile company name: {profile.company_name}", flush=True)
            
            # Update processing stage
            import copy
            profile_data = copy.deepcopy(profile.profile_data) if profile.profile_data else {}
            profile_data['processing_stage'] = 're_scraping_charika'
            profile.status = 'processing'
            profile.profile_data = profile_data
            db.session.commit()
            
            # Get company data using the custom Charika URL
            from services.charika_web import get_company_info_from_charika_url
            
            print(f"🔗 EXECUTING DIRECT CHARIKA SCRAPING - NO SEARCH", flush=True)
            print(f"🔗 Calling get_company_info_from_charika_url with URL: {charika_url}", flush=True)
            charika_data = get_company_info_from_charika_url(charika_url)
            print(f"🔗 Direct Charika scraping completed, got {len(charika_data) if charika_data else 0} fields", flush=True)
            
            if charika_data:
                # Merge with existing web data if it exists
                existing_web_data = profile_data.get('web_data', {})
                existing_basic_info = existing_web_data.get('basic_info', {})
                
                # Update contact information with Charika data
                if charika_data.get('contact'):
                    charika_contact = charika_data['contact']
                    existing_contact = existing_basic_info.get('contact', {})
                    
                    # Merge contact info, preferring Charika data
                    merged_contact = {
                        **existing_contact,
                        **{k: v for k, v in charika_contact.items() if v}  # Only add non-empty values
                    }
                    existing_basic_info['contact'] = merged_contact
                
                # Store the Charika URL in the basic_info
                existing_basic_info['charika_url'] = charika_url
                
                # Update the web data structure
                new_web_data = {
                    'basic_info': existing_basic_info,
                    **{k: v for k, v in existing_web_data.items() if k != 'basic_info'}
                }
                
                # Create a new profile_data dict instead of modifying the existing one
                updated_profile_data = {
                    **profile_data,
                    'web_data': new_web_data,
                    'processing_stage': 'kpis_extracted'
                }
                
                profile.profile_data = updated_profile_data
                profile.status = 'completed'
                db.session.commit()
                
                print(f"✅ CHARIKA RE-SCRAPING COMPLETE: Successfully extracted Charika data from URL: {charika_url}", flush=True)
                
                # Automatically trigger report generation after successful re-scraping
                print(f"🚀 Auto-triggering report generation for {profile.company_name}", flush=True)
                report_thread = threading.Thread(
                    target=process_final_analysis,
                    args=(app, db, CompanyProfile, profile_id, None),
                    daemon=True
                )
                report_thread.start()
                
            else:
                print(f"❌ Charika re-scraping failed for {profile.company_name}", flush=True)
                updated_profile_data = {
                    **profile_data,
                    'processing_stage': 'error'
                }
                profile.profile_data = updated_profile_data
                profile.status = 'failed'
                db.session.commit()
            
        except Exception as e:
            print(f"❌ Error in Charika re-scraping: {e}", flush=True)
            import traceback
            print(traceback.format_exc(), flush=True)
            try:
                profile = db.session.get(CompanyProfile, profile_id)
                if profile:
                    existing_profile_data = profile.profile_data or {}
                    updated_profile_data = {
                        **existing_profile_data,
                        'processing_stage': 'error'
                    }
                    profile.profile_data = updated_profile_data
                    profile.status = 'failed'
                    db.session.commit()
            except:
                pass
        finally:
            thread_tracker.remove_thread(profile_id, current_thread)

@app.route('/api/profiles', methods=['POST'])
@jwt_required()
def create_profile():
    """Create a new company profile. Rejects 'company name placeholder' as it's only for display purposes."""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        email_report = bool(data.get('email_report'))

        # Get company name - allow placeholder (will be extracted later in KPI page)
        company_name = data.get('company_name', '').strip()
        if not company_name:
            company_name = 'company name placeholder'  # Default if not provided

        # Convert fiscal year to string if provided
        fiscal_year = data.get('fiscal_year')
        if fiscal_year is not None:
            fiscal_year = str(fiscal_year)

        profile = CompanyProfile(
            company_name=company_name,
            fiscal_years=fiscal_year,  # Include fiscal year from frontend
            created_by=user_id,
            profile_data={
                'email_report': email_report
            }
        )
        
        db.session.add(profile)
        db.session.commit()
        
        # If user opted for email delivery, send a confirmation email now.
        # The actual PDF report will be sent when the profile is completed.
        if email_report:
            user = db.session.get(User, user_id)
            if user and user.email:
                send_email(
                    to_email=user.email,
                    subject='Report delivery preference confirmed',
                    body=(
                        'Hello,\n\n'
                        'You selected to receive the company analysis report by email.\n'
                        'We will automatically send you the PDF report as soon as your analysis is complete.\n\n'
                        'Best regards,\nCompany Profile Agent'
                    )
                )

        # Only start document processing for KPI extraction
        # News retrieval and web exploration will happen after KPI confirmation
        thread = threading.Thread(
            target=process_doc_processing,
            args=(app, db, CompanyProfile, LiasseDocument, profile.id),
            daemon=False
        )
        thread.start()

        return jsonify({
            'id': profile.id,
            'message': 'Profile created successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>', methods=['GET'])
@jwt_required()
def get_profile(profile_id):
    """Return full profile details including profile_data to diagnose failures."""
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)
        return jsonify({
            'id': profile.id,
            'company_name': profile.company_name,
            'status': profile.status,
            'profile_data': profile.profile_data or {},
            'created_at': profile.created_at.isoformat(),
            'updated_at': profile.updated_at.isoformat() if profile.updated_at else None
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/upload', methods=['POST'])
@jwt_required()
def upload_documents(profile_id):
    try:
        # Check if profile exists
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        # Check file count
        files = request.files.getlist('files')
        if len(files) > 3:
            return jsonify({'error': 'Maximum 3 files allowed'}), 400
        
        # For upload to existing profiles, we'll do a lightweight company name check
        # to avoid duplicate API calls that were already done during verification
        from services.profile_verification import extract_company_info_from_first_page, compare_company_names
        
        api_key = app.config.get('ANTHROPIC_API_KEY')
        document_company_names = []
        temp_file_paths = []
        
        try:
            # Save files temporarily for company name extraction
            import tempfile
            for i, file in enumerate(files):
                if file.filename == '':
                    continue
                    
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                    file.save(temp_file.name)
                    temp_file_paths.append(temp_file.name)
            
            # Only extract company names if we have multiple files or need verification
            # For single file uploads to existing profiles, skip the expensive extraction
            if len(temp_file_paths) > 1 or app.config.get('ENABLE_UPLOAD_VERIFICATION', False):
                print(f"[UPLOAD] Extracting company info from {len(temp_file_paths)} documents for comparison", flush=True)
                for i, file_path in enumerate(temp_file_paths):
                    company_info = extract_company_info_from_first_page(file_path, api_key)
                    if company_info and company_info.get('company_name'):
                        document_company_names.append(company_info['company_name'])
                        print(f"[UPLOAD] Document {i+1} company: {company_info['company_name']}", flush=True)
                    else:
                        print(f"[UPLOAD] Failed to extract company name from document {i+1}", flush=True)
                
                # Compare company names
                if document_company_names:
                    comparison_result = compare_company_names(profile.company_name, document_company_names)
                    
                    # If confirmation is required, return early with comparison details
                    if comparison_result['requires_confirmation']:
                        return jsonify({
                            'requires_confirmation': True,
                            'comparison_result': comparison_result,
                            'message': 'Company name mismatch detected - requires user confirmation',
                            'profile_company': profile.company_name,
                            'document_companies': document_company_names
                        }), 200
                else:
                    print(f"[UPLOAD] No company names extracted from documents, proceeding with upload", flush=True)
            else:
                print(f"[UPLOAD] Skipping company name extraction for single file upload to existing profile", flush=True)
            
            # If we reach here, company names match or comparison failed - proceed with upload
            uploaded_files = []
            
            for i, file in enumerate(files):
                if file.filename == '':
                    continue
                    
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
                
                # Copy from temp file to final location
                import shutil
                shutil.copy2(temp_file_paths[i], file_path)
                
                # Save document record
                document = LiasseDocument(
                    profile_id=profile_id,
                    file_name=filename,
                    file_path=file_path,
                    file_size=os.path.getsize(file_path)
                )
                
                db.session.add(document)
                uploaded_files.append({
                    'id': document.id,
                    'filename': filename,
                    'size': document.file_size
                })
            
            db.session.commit()
            
            return jsonify({
                'message': 'Files uploaded successfully',
                'files': uploaded_files,
                'requires_confirmation': False,
                'company_names_match': True
            })
            
        finally:
            # Clean up temporary files
            for temp_path in temp_file_paths:
                try:
                    os.unlink(temp_path)
                    print(f"[UPLOAD] Cleaned up temporary file: {temp_path}", flush=True)
                except:
                    pass
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/smart-upload', methods=['POST'])
@jwt_required()
def smart_upload_documents(profile_id):
    """
    Smart document upload that only processes new documents and reuses existing document data.
    """
    try:
        from services.profile_verification import identify_new_vs_existing_documents
        
        # Check if profile exists
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        # Check file count
        files = request.files.getlist('files')
        if len(files) > 3:
            return jsonify({'error': 'Maximum 3 files allowed'}), 400
        
        # Save files temporarily for analysis
        import tempfile
        temp_file_paths = []
        try:
            for i, file in enumerate(files):
                if file.filename == '':
                    continue
                    
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                    file.save(temp_file.name)
                    temp_file_paths.append(temp_file.name)
                    
            
            # Identify which documents are new vs existing
            # We need to extract company info first to pass to the function
            from services.profile_verification import extract_company_info_from_first_page
            
            # Extract company info from all documents to get the all_company_info parameter
            all_company_info = []
            api_key = app.config.get('ANTHROPIC_API_KEY')
            
            # Only extract company info if we have multiple files or if verification is explicitly enabled
            # For single file uploads after verification, we can skip the expensive extraction
            if len(temp_file_paths) > 1 or app.config.get('ENABLE_SMART_UPLOAD_VERIFICATION', False):
                print(f"[SMART_UPLOAD] Extracting company info from {len(temp_file_paths)} documents for analysis", flush=True)
                for i, file_path in enumerate(temp_file_paths):
                    company_info = extract_company_info_from_first_page(file_path, api_key)
                    if company_info:
                        all_company_info.append(company_info)
                        print(f"[SMART_UPLOAD] Document {i+1} extracted: {company_info}", flush=True)
                    else:
                        print(f"[SMART_UPLOAD] Failed to extract info from document {i+1}", flush=True)
            else:
                print(f"[SMART_UPLOAD] Skipping company info extraction for single file upload after verification", flush=True)
                # Create minimal company info for single file uploads
                for i, file_path in enumerate(temp_file_paths):
                    all_company_info.append({
                        'company_name': profile.company_name,  # Use profile company name
                        'fiscal_year': None  # Will be determined during processing
                    })
            
            # Now call the function with all required parameters
            document_analysis = identify_new_vs_existing_documents(
                db, CompanyProfile, temp_file_paths, profile.company_name, all_company_info
            )
            
            print(f"[SMART_UPLOAD] Document analysis: {document_analysis['total_new']} new, {document_analysis['total_existing']} existing", flush=True)
            
            # Check for company name mismatches
            from services.profile_verification import compare_company_names
            
            document_company_names = [info.get('company_name') for info in all_company_info if info.get('company_name')]
            if document_company_names:
                comparison_result = compare_company_names(profile.company_name, document_company_names)
                # print(f"[SMART_UPLOAD] Company name comparison result: {comparison_result}", flush=True)
                
                # If confirmation is required, return early with comparison details
                if comparison_result['requires_confirmation']:
                    return jsonify({
                        'requires_confirmation': True,
                        'comparison_result': comparison_result,
                        'message': 'Company name mismatch detected - requires user confirmation',
                        'profile_company': profile.company_name,
                        'document_companies': document_company_names,
                        'document_analysis': document_analysis
                    }), 200
            else:
                print(f"[SMART_UPLOAD] No company names extracted from documents, proceeding with upload", flush=True)
            
            uploaded_files = []
            processed_count = 0
            
            # Handle existing documents - reuse saved data
            for match in document_analysis['existing_matches']:
                print(f"[SMART_UPLOAD] Reusing existing document data for: {os.path.basename(match['file_path'])}", flush=True)
                
                # Save document record with existing data
                filename = os.path.basename(match['file_path'])
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
                
                # Copy the file to the new profile's directory
                import shutil
                shutil.copy2(match['file_path'], file_path)
                
                document = LiasseDocument(
                    profile_id=profile_id,
                    file_name=filename,
                    file_path=file_path,
                    file_size=os.path.getsize(file_path),
                    extracted_data=match['existing_data'].get('extracted_data'),
                    upload_status='reused',
                    ocr_status='completed'  # Mark as completed since we're reusing data
                )
                
                db.session.add(document)
                uploaded_files.append({
                    'id': document.id,
                    'filename': filename,
                    'size': document.file_size,
                    'status': 'reused',
                    'message': 'Document data reused from existing profile'
                })
            
            # Handle new documents - process normally
            for new_doc in document_analysis['new_documents']:
                # print(f"[SMART_UPLOAD] Processing new document: {os.path.basename(new_doc['file_path'])}", flush=True)
                
                filename = os.path.basename(new_doc['file_path'])
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
                
                # Copy the file to the new profile's directory
                import shutil
                shutil.copy2(new_doc['file_path'], file_path)
                
                # Save document record for new document
                document = LiasseDocument(
                    profile_id=profile_id,
                    file_name=filename,
                    file_path=file_path,
                    file_size=os.path.getsize(file_path),
                    upload_status='uploaded',
                    ocr_status='pending'  # Will be processed by the document processing pipeline
                )
                
                db.session.add(document)
                uploaded_files.append({
                    'id': document.id,
                    'filename': filename,
                    'size': document.file_size,
                    'status': 'new',
                    'message': 'Document will be processed'
                })
                processed_count += 1
            
            # Ensure all documents have the correct status for processing pipeline
            # print(f"[SMART_UPLOAD] Setting up documents for processing pipeline", flush=True)
            for doc in uploaded_files:
                if doc['status'] == 'new':
                    print(f"[SMART_UPLOAD] Document {doc['filename']} marked for processing (upload_status: uploaded, ocr_status: pending)", flush=True)
                elif doc['status'] == 'reused':
                    print(f"[SMART_UPLOAD] Document {doc['filename']} marked as reused (upload_status: reused, ocr_status: completed)", flush=True)
            
            # print(f"[SMART_UPLOAD] About to commit {len(uploaded_files)} documents to database", flush=True)
            db.session.commit()
            # print(f"[SMART_UPLOAD] Database commit completed successfully", flush=True)
            
            # Double-check that documents are actually in the database
            try:
                final_check = db.session.execute(
                    text("SELECT COUNT(*) as doc_count FROM liasse_documents WHERE profile_id = :profile_id"),
                    {"profile_id": profile_id}
                ).scalar()
                # print(f"[SMART_UPLOAD] Final database check: {final_check} documents found for profile {profile_id}", flush=True)
                
                if final_check != len(uploaded_files):
                    print(f"[SMART_UPLOAD] ⚠️ WARNING: Expected {len(uploaded_files)} documents but found {final_check} in database", flush=True)
                else:
                    print(f"[SMART_UPLOAD] ✅ SUCCESS: All {final_check} documents properly saved to database", flush=True)
                    
            except Exception as e:
                print(f"[SMART_UPLOAD] Error in final database check: {e}", flush=True)
            
            # Add a small delay to ensure database commit is fully processed
            import time
            time.sleep(1)
            # print(f"[SMART_UPLOAD] Added 1 second delay to ensure database synchronization", flush=True)
            
            # Start processing pipeline if there are new documents to process
            if processed_count > 0:
                print(f"[SMART_UPLOAD] Starting processing pipeline for {processed_count} new documents", flush=True)
                
                # Start web exploring directly (without news)
                web_thread = threading.Thread(
                    target=process_web_exploring,
                    args=(app, db, CompanyProfile, profile_id,),
                    daemon=False
                )
                web_thread.start()
                
                # Start doc processing (waits for web exploring to complete)
                doc_thread = threading.Thread(
                    target=process_doc_processing_with_web_sync,
                    args=(app, db, CompanyProfile, LiasseDocument, profile_id, web_thread),
                    daemon=False
                )
                doc_thread.start()
                
                print(f"[SMART_UPLOAD] Processing pipeline started for profile {profile_id} (web exploring + KPIs extraction)", flush=True)
            
            return jsonify({
                'message': 'Smart upload completed successfully',
                'uploaded_files': uploaded_files,
                'document_analysis': document_analysis,
                'new_documents_to_process': processed_count,
                'processing_started': processed_count > 0
            })
            
        finally:
            # Clean up temporary files
            for temp_path in temp_file_paths:
                try:
                    os.unlink(temp_path)
                    print(f"[SMART_UPLOAD] Cleaned up temporary file: {temp_path}", flush=True)
                except:
                    pass
        
    except Exception as e:
        db.session.rollback()
        print(f"[SMART_UPLOAD] Error in smart upload: {str(e)}", flush=True)
        return jsonify({'error': f'Smart upload failed: {str(e)}'}), 500

@app.route('/api/profiles/<profile_id>', methods=['DELETE'])
@jwt_required()
def delete_profile(profile_id):
    """Delete a profile and all associated documents"""
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        # Get the current user to check permissions
        user_id = get_jwt_identity()
        
        # Check if user is admin or the profile creator
        user = db.session.get(User, user_id)
        if user.role != 'admin' and profile.created_by != user_id:
            return jsonify({'error': 'Unauthorized to delete this profile'}), 403
        
        # Delete associated documents from filesystem (including markdown files)
        from services.doc_processing import cleanup_profile_files
        cleanup_profile_files(profile_id, app.config['UPLOAD_FOLDER'], db, LiasseDocument)
        
        # Delete the profile (CASCADE will delete associated documents from DB)
        db.session.delete(profile)
        db.session.commit()
        
        return jsonify({'message': 'Profile deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/complete', methods=['POST'])
@jwt_required()
def mark_profile_complete(profile_id):
    """Mark a profile as completed and optionally email the user a report link.

    Request JSON body (all optional):
      - report_url: string URL where the report can be downloaded/viewed
      - message: additional text to include in the email body
    """
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)

        # Only the creator or admins should be able to mark complete in a real app.
        # For now, require authentication and proceed.
        data = request.get_json(silent=True) or {}
        report_url = data.get('report_url')
        extra_message = data.get('message')

        # Update status and persist report_url into profile_data if provided
        profile.status = 'completed'
        current_data = profile.profile_data or {}
        if report_url:
            current_data['report_url'] = report_url
        profile.profile_data = current_data
        db.session.commit()

        # Email user if they opted in - send PDF report automatically
        email_report = bool((profile.profile_data or {}).get('email_report'))
        if email_report:
            # Send PDF report via email in a separate thread to avoid blocking
            import threading
            email_thread = threading.Thread(
                target=send_pdf_report_email,
                args=(profile_id,),
                daemon=True
            )
            email_thread.start()
            print(f"🚀 Started email thread for profile {profile_id}", flush=True)

        return jsonify({'message': 'Profile marked as completed'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/reprocess', methods=['POST'])
@jwt_required()
def reprocess_profile(profile_id):
    """Re-trigger background document processing for a profile."""
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)
        # Reset status to processing
        profile.status = 'processing'
        db.session.commit()

        # Start news retrieval first
        news_thread = threading.Thread(
            target=process_news_retrieval,
            args=(app, db, CompanyProfile, profile_id,),
            daemon=False
        )
        news_thread.start()
        
        # Start web exploring (waits for news to complete)
        web_thread = threading.Thread(
            target=process_web_exploring_with_news_sync,
            args=(app, db, CompanyProfile, profile_id, news_thread),
            daemon=False
        )
        web_thread.start()
        
        # Start doc processing (waits for web exploring to complete)
        thread = threading.Thread(
            target=process_doc_processing_with_web_sync,
            args=(app, db, CompanyProfile, LiasseDocument, profile_id, web_thread),
            daemon=False
        )
        thread.start()
        print(f"Started reprocessing thread for profile {profile_id}", flush=True)

        return jsonify({'message': 'Reprocessing started'}), 202
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/report', methods=['GET'])
def get_profile_report(profile_id):
    """Serve the generated HTML report for a completed profile"""
    try:
        # TODO: Re-enable authentication after testing
        # For now, let's test without auth to see if the report generation works
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        if profile.status != 'completed':
            return jsonify({'error': 'Profile is not completed yet'}), 400
            
        profile_data = profile.profile_data or {}
        extracted_kpis = profile_data.get('extracted_kpis')
        computed_ratios = profile_data.get('computed_ratios')
        web_data = profile_data.get('web_data', {})
        
        # Always use profile.company_name first (it's the source of truth and gets updated by re-scraping)
        # Check if profile_data has a placeholder, and if so, use profile.company_name instead
        company_name_from_data = profile_data.get('company_name')
        placeholder_values = ['company name placeholder', 'COMPANY NAME PLACEHOLDER', 
                             'Company Name Not Extracted', 'COMPANY NAME NOT EXTRACTED']
        
        if company_name_from_data and company_name_from_data.strip() in placeholder_values:
            # profile_data has placeholder, use profile.company_name (which is updated by re-scraping)
            company_name = profile.company_name
            print(f"📝 Report: Using profile.company_name '{company_name}' instead of placeholder '{company_name_from_data}'", flush=True)
        else:
            # Use profile_data company_name if not a placeholder, otherwise fall back to profile.company_name
            company_name = company_name_from_data or profile.company_name
        
        # Clean the company name for display by removing legal forms
        from services.profile_verification import _normalize_company_name
        display_company_name = _normalize_company_name(company_name)
        
        # Final check: if display_company_name is still a placeholder after normalization, use profile.company_name
        normalized_placeholder_check = display_company_name.upper().strip()
        if normalized_placeholder_check in ['COMPANY NAME PLACEHOLDER', 'COMPANY NAME NOT EXTRACTED']:
            display_company_name = _normalize_company_name(profile.company_name)
            print(f"📝 Report: Replaced placeholder '{normalized_placeholder_check}' with profile.company_name: '{display_company_name}'", flush=True)
        
        # Debug: Log the profile data structure (commented out for production)
        # print(f"🔍 DEBUG: Profile data keys: {list(profile_data.keys()) if profile_data else 'None'}", flush=True)
        # print(f"🔍 DEBUG: extracted_kpis type: {type(extracted_kpis)}, keys: {list(extracted_kpis.keys()) if extracted_kpis else 'None'}", flush=True)
        # print(f"🔍 DEBUG: computed_ratios type: {type(computed_ratios)}, keys: {list(computed_ratios.keys()) if computed_ratios else 'None'}", flush=True)
        # print(f"🔍 DEBUG: web_data keys: {list(web_data.keys()) if web_data else 'None'}", flush=True)
        # if web_data:
        #     print(f"🔍 DEBUG: recommendation length: {len(web_data.get('recommendation', ''))}", flush=True)
        #     print(f"🔍 DEBUG: detailed_analysis length: {len(web_data.get('detailed_analysis', ''))}", flush=True)
        
        if not extracted_kpis and not computed_ratios:
            print(f"❌ ERROR: No financial data found in profile {profile_id}", flush=True)
            print(f"❌ Profile status: {profile.status}", flush=True)
            print(f"❌ Profile data keys: {list(profile_data.keys()) if profile_data else 'None'}", flush=True)
            
            # Provide fallback test data for debugging
            print(f"🔄 Providing fallback test data for debugging...", flush=True)
            extracted_kpis = {
                'Chiffre d\'affaires': {'N': 1000000, 'N-1': 900000},
                'Résultat Net': {'N': 100000, 'N-1': 80000},
                'Capitaux propres': {'N': 500000, 'N-1': 450000},
                'Dettes de financement': {'N': 200000, 'N-1': 180000},
                'Trésorerie-Actif': {'N': 50000, 'N-1': 40000}
            }
            computed_ratios = {
                'marge_nette_n': 10.0,
                'marge_exploitation_n': 15.0,
                'roe_n': 20.0,
                'roce_n': 25.0,
                'gearing_n': 30.0,
                'marge_ebitda_n': 18.0
            }
            print(f"✅ Fallback data created: {len(extracted_kpis)} KPIs, {len(computed_ratios)} ratios", flush=True)
        
        # Read the report template - now that frontend is mounted at /app/frontend
        template_path = os.path.join('frontend', 'src', 'report_template', 'report.html')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Report template not found'}), 500
            
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
            
        # Read main.js content
        main_js_path = os.path.join('frontend', 'src', 'js', 'main.js')
        if os.path.exists(main_js_path):
            with open(main_js_path, 'r', encoding='utf-8') as f:
                main_js_content = f.read()
        else:
            main_js_content = "console.error('main.js not found');"
            
        # Read style.css content
        style_css_path = os.path.join('frontend', 'src', 'css', 'style.css')
        if os.path.exists(style_css_path):
            with open(style_css_path, 'r', encoding='utf-8') as f:
                style_css_content = f.read()
        else:
            style_css_content = "/* style.css not found */"
            
        # Create the data structure for the template
        basic_info = web_data.get('basic_info', {})
        company_overview = basic_info.get('companyOverview', {})
        
        # Get analysis sections from profile data (generated by financial_reporting.py)
        profile_data = profile.profile_data or {}
        
        template_data = {
            'company_name': company_name,
            'extracted_kpis': extracted_kpis or {},
            'computed_ratios': computed_ratios or {},
            'fiscal_years': profile.fiscal_years,  # Add fiscal years information
            # Company overview data from web exploring
            'companyOverview': {
                'companyFoundationyear': company_overview.get('companyFoundationyear', 'Non spécifié'),
                'companyExpertise': company_overview.get('companyExpertise', 'À déterminer'),
                'primary_sector': company_overview.get('primary_sector', 'Secteur général'),
                'legal_form': company_overview.get('legal_form', 'SARL'),
                'companyDefinition': company_overview.get('companyDefinition', f'Entreprise {display_company_name}'),
                'staff_count': company_overview.get('staff_count', 'À préciser')
            },
            # Sectors and markets - handle null values safely
            'sectors': basic_info.get('sectors') if basic_info.get('sectors') is not None else [],
            'markets': basic_info.get('markets') if basic_info.get('markets') is not None else [],
            'keyPeople': basic_info.get('keyPeople') if basic_info.get('keyPeople') is not None else [],
            # Contact information
            'contact': basic_info.get('contact', {
                'phone': 'Non disponible',
                'email': 'Non disponible',
                'address': 'Adresse à préciser',
                'website': 'Non disponible'
            }),
            # Analysis sections from financial_reporting.py
            'recommendation': profile_data.get('recommendation', 'Recommandation à définir'),
            'news': profile_data.get('news_data', {}).get('analysis', 'Actualités sectorielles à rechercher') if profile_data.get('news_data') else 'Actualités sectorielles à rechercher',
            'news_urls': profile_data.get('news_data', {}).get('urls', []) if profile_data.get('news_data') else [],
            'news_articles': profile_data.get('news_data', {}).get('urls', []) if profile_data.get('news_data') else [],
            'detailed_analysis': profile_data.get('detailed_analysis', 'Analyse détaillée à compléter'),
            # SWOT analysis data from financial_reporting.py
            'swot_analysis': profile_data.get('swot_analysis', {
                'strengths': [],
                'weaknesses': [],
                'opportunities': [],
                'threats': []
            })
        }
        
        # Debug: Log the data being passed to the template
        # print(f"🔍 DEBUG: Template data structure:")
        # print(f"🔍 - company_name: {template_data['company_name']}")
        # print(f"🔍 - extracted_kpis keys: {list(template_data['extracted_kpis'].keys())}")
        # print(f"🔍 - computed_ratios keys: {list(template_data['computed_ratios'].keys())}")
        # print(f"🔍 - extracted_kpis sample: {dict(list(template_data['extracted_kpis'].items())[:3]) if template_data['extracted_kpis'] else 'Empty'}")
        # print(f"🔍 - computed_ratios sample: {dict(list(template_data['computed_ratios'].items())[:3]) if template_data['computed_ratios'] else 'Empty'}")
        
        # Additional debugging for data structure
        if template_data['extracted_kpis']:
            # print(f"🔍 DEBUG: First KPI structure:")
            first_kpi_key = list(template_data['extracted_kpis'].keys())[0]
            first_kpi_value = template_data['extracted_kpis'][first_kpi_key]
            # print(f"🔍 - Key: {first_kpi_key}")
            # print(f"🔍 - Value type: {type(first_kpi_value)}")
            # print(f"🔍 - Value: {first_kpi_value}")
        
        if template_data['computed_ratios']:
            # print(f"🔍 DEBUG: First ratio structure:")
            first_ratio_key = list(template_data['computed_ratios'].keys())[0]
            first_ratio_value = template_data['computed_ratios'][first_ratio_key]
            # print(f"🔍 - Key: {first_ratio_key}")
            # print(f"🔍 - Value type: {type(first_ratio_value)}")
            # print(f"🔍 - Value: {first_ratio_value}")
        
        # Inject the data and main.js directly into the HTML
        template_data_json = json.dumps(template_data, ensure_ascii=False, indent=2)
        
        # Replace template placeholders
        html_content = template_content.replace(
            '<script id="report-data" type="application/json">\n  {{ reportData | tojson }}\n</script>',
            f'<script id="report-data" type="application/json">\n{template_data_json}\n</script>'
        )
        
        html_content = html_content.replace(
            '<script src="{{ url_for(\'static\', filename=\'js/main.js\') }}"></script>',
            f'<script>\n{main_js_content}\n</script>'
        )
        
        # Replace CSS link with actual CSS content
        html_content = html_content.replace(
            '<link rel="stylesheet" href="{{ url_for(\'static\', filename=\'css/style.css\') }}">',
            f'<style>\n{style_css_content}\n</style>'
        )
        
        # Generate sectors HTML
        sectors_html = ""
        sectors = basic_info.get('sectors')
        if sectors and isinstance(sectors, list):
            for sector in sectors:
                if sector and isinstance(sector, dict):
                    sectors_html += f'''
                    <div class="info-card">
                        <div class="icon"><i class="fas fa-signal"></i></div>
                        <h4>{sector.get('title', 'Secteur')}</h4>
                        <p>{sector.get('description', 'Description non disponible')}</p>
                    </div>'''
        
        # Generate markets HTML
        markets_html = ""
        markets = basic_info.get('markets')
        if markets and isinstance(markets, list):
            for market in markets:
                if market and isinstance(market, dict):
                    markets_html += f'''
                    <div class="info-card">
                        <div class="icon"><i class="fas fa-landmark"></i></div>
                        <h4>{market.get('title', 'Marché')}</h4>
                        <p>{market.get('description', 'Description non disponible')}</p>
                    </div>'''
        
        # Helper function to safely get string values (handle None values)
        def safe_get(value, default=''):
            return str(value) if value is not None else default
        
        # Generate key people HTML for standalone section
        key_people_html = ""
        key_people = basic_info.get('keyPeople')
        if key_people and isinstance(key_people, list):
            for person in key_people:
                if person and isinstance(person, dict):
                    # Skip people with all None values
                    if not any([person.get('name'), person.get('position'), person.get('initials')]):
                        continue
                    key_people_html += f'''
                    <div class="person-item">
                        <div class="person-avatar">{safe_get(person.get('initials'), 'N/A')}</div>
                        <div class="person-info">
                            <h4>{safe_get(person.get('name'), 'Nom non disponible')}</h4>
                            <p>{safe_get(person.get('position'), 'Poste non spécifié')}</p>
                        </div>
                    </div>'''
        
        # Generate compact key people HTML for header-overview section
        key_people_compact_html = ""
        key_people_list = basic_info.get('keyPeople')
        if key_people_list and isinstance(key_people_list, list):
            for i, person in enumerate(key_people_list):
                if person and isinstance(person, dict):
                    # Skip people with all None values
                    if not any([person.get('name'), person.get('position')]):
                        continue
                    
                    # Start new line every 2 people
                    if i % 2 == 0:
                        if i > 0:  # Close previous line
                            key_people_compact_html += '</div>'
                        key_people_compact_html += '<div class="key-people-row">'
                    
                    key_people_compact_html += f'<span class="key-person-item">• {safe_get(person.get("name"), "Nom non disponible")}</span>'
        
        # Close the last row if there are any people
        if key_people_list:
            key_people_compact_html += '</div>'
        
        # Replace all Jinja2 template variables with actual data
        html_content = html_content.replace('{{ reportData.companyName }}', display_company_name)
        html_content = html_content.replace('{{ reportData.companyOverview.companyExpertise }}', safe_get(company_overview.get('companyExpertise'), 'À déterminer'))
        html_content = html_content.replace('{{ reportData.companyOverview.companyDefinition }}', safe_get(company_overview.get('companyDefinition'), f'Entreprise {display_company_name}'))
        html_content = html_content.replace('{{ reportData.companyOverview.primary_sector }}', safe_get(company_overview.get('primary_sector'), 'Secteur général'))
        html_content = html_content.replace('{{ reportData.companyOverview.legal_form }}', safe_get(company_overview.get('legal_form'), 'SARL'))
        html_content = html_content.replace('{{ reportData.companyOverview.companyFoundationyear }}', safe_get(company_overview.get('companyFoundationyear'), 'Non spécifié'))
        html_content = html_content.replace('{{ reportData.companyOverview.staff_count }}', safe_get(company_overview.get('staff_count'), 'À préciser'))
        html_content = html_content.replace('{{ reportData.contact.address }}', safe_get(basic_info.get('contact', {}).get('address'), 'Adresse à préciser'))
        html_content = html_content.replace('{{ reportData.contact.phone }}', safe_get(basic_info.get('contact', {}).get('phone'), 'Non disponible'))
        html_content = html_content.replace('{{ reportData.contact.email }}', safe_get(basic_info.get('contact', {}).get('email'), 'Non disponible'))
        html_content = html_content.replace('{{ reportData.contact.website }}', safe_get(basic_info.get('contact', {}).get('website'), 'Non disponible'))
        
        # Generate news content with links
        def generate_news_html(news_text, news_articles):
            import html
            news_html = f'<p>{html.escape(news_text) if news_text else "Actualités sectorielles à rechercher"}</p>'
            
            return news_html

        # Debug: Log the textual data before replacement
        # Get analysis data from profile_data (financial_reporting.py) instead of web_data
        recommendation_text = profile_data.get('recommendation', 'Recommandation à définir')
        news_data = profile_data.get('news_data', {})
        news_text = news_data.get('analysis', 'Actualités sectorielles à rechercher') if news_data else 'Actualités sectorielles à rechercher'
        news_articles = news_data.get('urls', []) if news_data else []
        detailed_analysis_text = profile_data.get('detailed_analysis', 'Analyse détaillée à compléter')
        
        # print(f"🔍 DEBUG: About to replace recommendation: {len(recommendation_text)} chars", flush=True)
        # print(f"🔍 DEBUG: About to replace news: {len(news_text)} chars with {len(news_articles)} articles", flush=True)
        # print(f"🔍 DEBUG: About to replace detailed_analysis: {len(detailed_analysis_text)} chars", flush=True)
        
        # Ensure proper encoding and escape HTML characters
        import html
        recommendation_text_safe = html.escape(recommendation_text) if recommendation_text else 'Recommandation à définir'
        news_html_content = generate_news_html(news_text, news_articles)
        detailed_analysis_text_safe = html.escape(detailed_analysis_text) if detailed_analysis_text else 'Analyse détaillée à compléter'
        
        html_content = html_content.replace('{{ reportData.recommendation }}', recommendation_text_safe)
        html_content = html_content.replace('{{ reportData.news }}', news_html_content)
        html_content = html_content.replace('{{ reportData.detailed_analysis }}', detailed_analysis_text_safe)
        
        # Replace loop sections
        html_content = html_content.replace('{% for sector in reportData.sectors %}\n            <div class="info-card">\n                <div class="icon"><i class="fas fa-signal"></i></div>\n                <h4>{{ sector.title }}</h4>\n                <p>{{ sector.description }}</p>\n            </div>\n            {% endfor %}', sectors_html)
        html_content = html_content.replace('{% for market in reportData.markets %}\n          <div class="info-card">\n            <div class="icon"><i class="fas fa-landmark"></i></div>\n            <h4>{{ market.title }}</h4>\n            <p>{{ market.description }}</p>\n          </div>\n          {% endfor %}', markets_html)
        html_content = html_content.replace('{% for person in reportData.keyPeople %}\n          <div class="person-item">\n            <div class="person-avatar">{{ person.initials }}</div>\n            <div class="person-info">\n              <h4>{{ person.name }}</h4>\n              <p>{{ person.position }}</p>\n            </div>\n          </div>\n          {% endfor %}', key_people_html)
        
        # Replace header-overview keyPeople loop section
        html_content = html_content.replace('<!-- Fourth row: Dirigeants (compact) -->\n          <div class="overview-row">\n            <div class="overview-bullet-item full-width">\n              <i class="fas fa-users" style="color: white; margin-right: 8px;"></i>\n              <span><strong>Dirigeants:</strong></span>\n              <div class="key-people-compact">\n                {% for person in reportData.keyPeople %}\n                <span class="key-person-compact">\n                  <span class="key-person-compact-avatar">{{ person.initials }}</span>\n                  {{ person.name }} ({{ person.position }})\n                </span>\n                {% if not loop.last %}<span class="separator"> • </span>{% endif %}\n                {% endfor %}\n              </div>\n            </div>\n          </div>', f'''<!-- Fourth row: Dirigeants (compact) -->
          <div class="overview-row">
            <div class="overview-bullet-item full-width">
              <i class="fas fa-users" style="color: white; margin-right: 8px;"></i>
              <span><strong>Dirigeants:</strong></span>
              <div class="key-people-compact">
                {key_people_compact_html}
              </div>
            </div>
          </div>''')
        
        # Replace company name placeholders
        html_content = html_content.replace('{{ company_name }}', company_name)
        html_content = html_content.replace('{{ companyName }}', display_company_name)
        
        # Remove any remaining template variables that should be handled by JavaScript
        html_content = html_content.replace('{{ reportData.financialData.metrics.gearing }}', '')
        
        # Remove remaining Jinja2 syntax - the JavaScript will handle data population
        import re
        html_content = re.sub(r'\{\%.*?\%\}', '', html_content, flags=re.DOTALL)
        key_people = basic_info.get('keyPeople') or []
        html_content = html_content.replace('{{ reportData.keyPeople|length }}', str(len(key_people)))
        html_content = re.sub(r'\{\{.*?\}\}', '', html_content)
        
        return html_content, 200, {'Content-Type': 'text/html; charset=utf-8'}
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/pdf', methods=['GET'])
@jwt_required()
def get_profile_pdf(profile_id):
    """Generate and serve a PDF version of the profile report"""
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        if profile.status != 'completed':
            return jsonify({'error': 'Profile is not completed yet'}), 400
            
        # Get the HTML content from the existing report endpoint logic
        # We'll reuse the same logic but generate PDF instead of returning HTML
        profile_data = profile.profile_data or {}
        extracted_kpis = profile_data.get('extracted_kpis')
        computed_ratios = profile_data.get('computed_ratios')
        web_data = profile_data.get('web_data', {})
        
        # Always use profile.company_name first (it's the source of truth and gets updated by re-scraping)
        # Check if profile_data has a placeholder, and if so, use profile.company_name instead
        company_name_from_data = profile_data.get('company_name')
        placeholder_values = ['company name placeholder', 'COMPANY NAME PLACEHOLDER', 
                             'Company Name Not Extracted', 'COMPANY NAME NOT EXTRACTED']
        
        if company_name_from_data and company_name_from_data.strip() in placeholder_values:
            # profile_data has placeholder, use profile.company_name (which is updated by re-scraping)
            company_name = profile.company_name
            print(f"📝 PDF Report: Using profile.company_name '{company_name}' instead of placeholder '{company_name_from_data}'", flush=True)
        else:
            # Use profile_data company_name if not a placeholder, otherwise fall back to profile.company_name
            company_name = company_name_from_data or profile.company_name
        
        # Clean the company name for display by removing legal forms
        from services.profile_verification import _normalize_company_name
        display_company_name = _normalize_company_name(company_name)
        
        # Final check: if display_company_name is still a placeholder after normalization, use profile.company_name
        normalized_placeholder_check = display_company_name.upper().strip()
        if normalized_placeholder_check in ['COMPANY NAME PLACEHOLDER', 'COMPANY NAME NOT EXTRACTED']:
            display_company_name = _normalize_company_name(profile.company_name)
            print(f"📝 PDF Report: Replaced placeholder '{normalized_placeholder_check}' with profile.company_name: '{display_company_name}'", flush=True)
        
        if not extracted_kpis and not computed_ratios:
            return jsonify({'error': 'Financial data not available'}), 404
            
        # Read the report template
        template_path = os.path.join('frontend', 'src', 'report_template', 'report.html')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Report template not found'}), 500
            
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
            
        # Read main.js content
        main_js_path = os.path.join('frontend', 'src', 'js', 'main.js')
        if os.path.exists(main_js_path):
            with open(main_js_path, 'r', encoding='utf-8') as f:
                main_js_content = f.read()
        else:
            main_js_content = "console.error('main.js not found');"
            
        # Read style.css content
        style_css_path = os.path.join('frontend', 'src', 'css', 'style.css')
        if os.path.exists(style_css_path):
            with open(style_css_path, 'r', encoding='utf-8') as f:
                style_css_content = f.read()
        else:
            style_css_content = "/* style.css not found */"
            
        # Create the data structure for the template
        basic_info = web_data.get('basic_info', {})
        company_overview = basic_info.get('companyOverview', {})
        
        # Get analysis sections from profile data (generated by financial_reporting.py)
        profile_data = profile.profile_data or {}
        
        template_data = {
            'company_name': company_name,
            'extracted_kpis': extracted_kpis or {},
            'computed_ratios': computed_ratios or {},
            'fiscal_years': profile.fiscal_years,  # Add fiscal years information
            # Company overview data from web exploring
            'companyOverview': {
                'companyFoundationyear': company_overview.get('companyFoundationyear', 'Non spécifié'),
                'companyExpertise': company_overview.get('companyExpertise', 'À déterminer'),
                'primary_sector': company_overview.get('primary_sector', 'Secteur général'),
                'legal_form': company_overview.get('legal_form', 'SARL'),
                'companyDefinition': company_overview.get('companyDefinition', f'Entreprise {display_company_name}'),
                'staff_count': company_overview.get('staff_count', 'À préciser')
            },
            # Sectors and markets - handle null values safely
            'sectors': basic_info.get('sectors') if basic_info.get('sectors') is not None else [],
            'markets': basic_info.get('markets') if basic_info.get('markets') is not None else [],
            'keyPeople': basic_info.get('keyPeople') if basic_info.get('keyPeople') is not None else [],
            # Contact information
            'contact': basic_info.get('contact', {
                'phone': 'Non disponible',
                'email': 'Non disponible',
                'address': 'Adresse à préciser',
                'website': 'Non disponible'
            }),
            # Analysis sections from financial_reporting.py
            'recommendation': profile_data.get('recommendation', 'Recommandation à définir'),
            'news': profile_data.get('news_data', {}).get('analysis', 'Actualités sectorielles à rechercher') if profile_data.get('news_data') else 'Actualités sectorielles à rechercher',
            'news_urls': profile_data.get('news_data', {}).get('urls', []) if profile_data.get('news_data') else [],
            'news_articles': profile_data.get('news_data', {}).get('urls', []) if profile_data.get('news_data') else [],
            'detailedAnalysis': profile_data.get('detailed_analysis', 'Analyse détaillée à compléter'),
            # SWOT analysis data from financial_reporting.py
            'swot_analysis': profile_data.get('swot_analysis', {
                'strengths': [],
                'weaknesses': [],
                'opportunities': [],
                'threats': []
            })
        }
        
        # Debug: Log the data being passed to the template
        # print(f"🔍 DEBUG: Template data structure:")
        # print(f"🔍 - company_name: {template_data['company_name']}")
        # print(f"🔍 - extracted_kpis keys: {list(template_data['extracted_kpis'].keys())}")
        # print(f"🔍 - computed_ratios keys: {list(template_data['computed_ratios'].keys())}")
        # print(f"🔍 - extracted_kpis sample: {dict(list(template_data['extracted_kpis'].items())[:3]) if template_data['extracted_kpis'] else 'Empty'}")
        # print(f"🔍 - computed_ratios sample: {dict(list(template_data['computed_ratios'].items())[:3]) if template_data['computed_ratios'] else 'Empty'}")
        
        # Additional debugging for data structure
        if template_data['extracted_kpis']:
            # print(f"🔍 DEBUG: First KPI structure:")
            first_kpi_key = list(template_data['extracted_kpis'].keys())[0]
            first_kpi_value = template_data['extracted_kpis'][first_kpi_key]
            # print(f"🔍 - Key: {first_kpi_key}")
            # print(f"🔍 - Value type: {type(first_kpi_value)}")
            # print(f"🔍 - Value: {first_kpi_value}")
        
        if template_data['computed_ratios']:
            # print(f"🔍 DEBUG: First ratio structure:")
            first_ratio_key = list(template_data['computed_ratios'].keys())[0]
            first_ratio_value = template_data['computed_ratios'][first_ratio_key]
            # print(f"🔍 - Key: {first_ratio_key}")
            # print(f"🔍 - Value type: {type(first_ratio_value)}")
            # print(f"🔍 - Value: {first_ratio_value}")
        
        # Inject the data and main.js directly into the HTML
        template_data_json = json.dumps(template_data, ensure_ascii=False, indent=2)
        
        # Replace template placeholders
        html_content = template_content.replace(
            '<script id="report-data" type="application/json">\n  {{ reportData | tojson }}\n</script>',
            f'<script id="report-data" type="application/json">\n{template_data_json}\n</script>'
        )
        
        html_content = html_content.replace(
            '<script src="{{ url_for(\'static\', filename=\'js/main.js\') }}"></script>',
            f'<script>\n{main_js_content}\n</script>'
        )
        
        # Replace CSS link with actual CSS content
        html_content = html_content.replace(
            '<link rel="stylesheet" href="{{ url_for(\'static\', filename=\'css/style.css\') }}">',
            f'<style>\n{style_css_content}\n</style>'
        )
        
        # Generate sectors HTML
        sectors_html = ""
        sectors = basic_info.get('sectors')
        if sectors and isinstance(sectors, list):
            for sector in sectors:
                if sector and isinstance(sector, dict):
                    sectors_html += f'''
                    <div class="info-card">
                        <div class="icon"><i class="fas fa-signal"></i></div>
                        <h4>{sector.get('title', 'Secteur')}</h4>
                        <p>{sector.get('description', 'Description non disponible')}</p>
                    </div>'''
        
        # Generate markets HTML
        markets_html = ""
        markets = basic_info.get('markets')
        if markets and isinstance(markets, list):
            for market in markets:
                if market and isinstance(market, dict):
                    markets_html += f'''
                    <div class="info-card">
                        <div class="icon"><i class="fas fa-landmark"></i></div>
                        <h4>{market.get('title', 'Marché')}</h4>
                        <p>{market.get('description', 'Description non disponible')}</p>
                    </div>'''
        
        # Helper function to safely get string values (handle None values)
        def safe_get(value, default=''):
            return str(value) if value is not None else default
        
        # Generate key people HTML for standalone section
        key_people_html = ""
        key_people = basic_info.get('keyPeople')
        if key_people and isinstance(key_people, list):
            for person in key_people:
                if person and isinstance(person, dict):
                    # Skip people with all None values
                    if not any([person.get('name'), person.get('position'), person.get('initials')]):
                        continue
                    key_people_html += f'''
                    <div class="person-item">
                        <div class="person-avatar">{safe_get(person.get('initials'), 'N/A')}</div>
                        <div class="person-info">
                            <h4>{safe_get(person.get('name'), 'Nom non disponible')}</h4>
                            <p>{safe_get(person.get('position'), 'Poste non spécifié')}</p>
                        </div>
                    </div>'''
        
        # Generate compact key people HTML for header-overview section
        key_people_compact_html = ""
        key_people_list = basic_info.get('keyPeople')
        if key_people_list and isinstance(key_people_list, list):
            for i, person in enumerate(key_people_list):
                if person and isinstance(person, dict):
                    # Skip people with all None values
                    if not any([person.get('name'), person.get('position')]):
                        continue
                    position = safe_get(person.get("position"), "Poste non spécifié")
                    # Remove "(Dirigeant)" from position if it exists
                    position = position.replace(" (Dirigeant)", "").replace("(Dirigeant)", "").replace("Dirigeant", "")
                    
                    # Start new line every 2 people
                    if i % 2 == 0:
                        if i > 0:  # Close previous line
                            key_people_compact_html += '</div>'
                        key_people_compact_html += '<div class="key-people-row">'
                    
                    key_people_compact_html += f'<span class="key-person-item">• {safe_get(person.get("name"), "Nom non disponible")} ({position})</span>'
        
        # Close the last row if there are any people
        if key_people_list:
            key_people_compact_html += '</div>'
        
        # Replace all Jinja2 template variables with actual data
        html_content = html_content.replace('{{ reportData.companyName }}', display_company_name)
        html_content = html_content.replace('{{ reportData.companyOverview.companyExpertise }}', safe_get(company_overview.get('companyExpertise'), 'À déterminer'))
        html_content = html_content.replace('{{ reportData.companyOverview.companyDefinition }}', safe_get(company_overview.get('companyDefinition'), f'Entreprise {display_company_name}'))
        html_content = html_content.replace('{{ reportData.companyOverview.primary_sector }}', safe_get(company_overview.get('primary_sector'), 'Secteur général'))
        html_content = html_content.replace('{{ reportData.companyOverview.legal_form }}', safe_get(company_overview.get('legal_form'), 'SARL'))
        html_content = html_content.replace('{{ reportData.companyOverview.companyFoundationyear }}', safe_get(company_overview.get('companyFoundationyear'), 'Non spécifié'))
        html_content = html_content.replace('{{ reportData.companyOverview.staff_count }}', safe_get(company_overview.get('staff_count'), 'À préciser'))
        html_content = html_content.replace('{{ reportData.contact.address }}', safe_get(basic_info.get('contact', {}).get('address'), 'Adresse à préciser'))
        html_content = html_content.replace('{{ reportData.contact.phone }}', safe_get(basic_info.get('contact', {}).get('phone'), 'Non disponible'))
        html_content = html_content.replace('{{ reportData.contact.email }}', safe_get(basic_info.get('contact', {}).get('email'), 'Non disponible'))
        html_content = html_content.replace('{{ reportData.contact.website }}', safe_get(basic_info.get('contact', {}).get('website'), 'Non disponible'))
        
        # Generate news content with links (same function as HTML report)
        def generate_news_html(news_text, news_articles):
            import html
            news_html = f'<p>{html.escape(news_text) if news_text else "Actualités sectorielles à rechercher"}</p>'
            
            return news_html

        # Debug: Log the textual data before replacement
        recommendation_text = web_data.get('recommendation', 'Recommandation à définir')
        news_text = web_data.get('news', 'Actualités sectorielles à rechercher')
        news_articles = web_data.get('news_articles', [])
        detailed_analysis_text = web_data.get('detailed_analysis', 'Analyse détaillée à compléter')
        
        # Ensure proper encoding and escape HTML characters
        import html
        recommendation_text_safe = html.escape(recommendation_text) if recommendation_text else 'Recommandation à définir'
        news_html_content = generate_news_html(news_text, news_articles)
        detailed_analysis_text_safe = html.escape(detailed_analysis_text) if detailed_analysis_text else 'Analyse détaillée à compléter'
        
        html_content = html_content.replace('{{ reportData.recommendation }}', recommendation_text_safe)
        html_content = html_content.replace('{{ reportData.news }}', news_html_content)
        html_content = html_content.replace('{{ reportData.detailedAnalysis }}', detailed_analysis_text_safe)
        
        # Replace loop sections
        html_content = html_content.replace('{% for sector in reportData.sectors %}\n            <div class="info-card">\n                <div class="icon"><i class="fas fa-signal"></i></div>\n                <h4>{{ sector.title }}</h4>\n                <p>{{ sector.description }}</p>\n            </div>\n            {% endfor %}', sectors_html)
        html_content = html_content.replace('{% for market in reportData.markets %}\n          <div class="info-card">\n            <div class="icon"><i class="fas fa-landmark"></i></div>\n            <h4>{{ market.title }}</h4>\n            <p>{{ market.description }}</p>\n          </div>\n          {% endfor %}', markets_html)
        html_content = html_content.replace('{% for person in reportData.keyPeople %}\n          <div class="person-item">\n            <div class="person-avatar">{{ person.initials }}</div>\n            <div class="person-info">\n              <h4>{{ person.name }}</h4>\n              <p>{{ person.position }}</p>\n            </div>\n          </div>\n          {% endfor %}', key_people_html)
        
        # Replace header-overview keyPeople loop section
        html_content = html_content.replace('<!-- Fourth row: Dirigeants (compact) -->\n          <div class="overview-row">\n            <div class="overview-bullet-item full-width">\n              <i class="fas fa-users" style="color: white; margin-right: 8px;"></i>\n              <span><strong>Dirigeants:</strong></span>\n              <div class="key-people-compact">\n                {% for person in reportData.keyPeople %}\n                <span class="key-person-compact">\n                  <span class="key-person-compact-avatar">{{ person.initials }}</span>\n                  {{ person.name }} ({{ person.position }})\n                </span>\n                {% if not loop.last %}<span class="separator"> • </span>{% endif %}\n                {% endfor %}\n              </div>\n            </div>\n          </div>', f'''<!-- Fourth row: Dirigeants (compact) -->
          <div class="overview-row">
            <div class="overview-bullet-item full-width">
              <i class="fas fa-users" style="color: white; margin-right: 8px;"></i>
              <span><strong>Dirigeants:</strong></span>
              <div class="key-people-compact">
                {key_people_compact_html}
              </div>
            </div>
          </div>''')
        
        # Replace company name placeholders
        html_content = html_content.replace('{{ company_name }}', company_name)
        html_content = html_content.replace('{{ companyName }}', display_company_name)
        
        # Remove any remaining template variables that should be handled by JavaScript
        html_content = html_content.replace('{{ reportData.financialData.metrics.gearing }}', '')
        
        # Remove remaining Jinja2 syntax - the JavaScript will handle data population
        import re
        html_content = re.sub(r'\{\%.*?\%\}', '', html_content, flags=re.DOTALL)
        key_people = basic_info.get('keyPeople') or []
        html_content = html_content.replace('{{ reportData.keyPeople|length }}', str(len(key_people)))
        html_content = re.sub(r'\{\{.*?\}\}', '', html_content)
        
        # Generate PDF using weasyprint with safer options
        from weasyprint import HTML, CSS
        from weasyprint.text.fonts import FontConfiguration
        from io import BytesIO
        
        # Create a BytesIO buffer to store the PDF
        pdf_buffer = BytesIO()
        
        # Create a simplified HTML version without external dependencies
        # Remove external CDN links that might cause recursion issues
        import re
        
        # Remove external script and link tags
        html_content = re.sub(r'<script[^>]*src="https?://[^"]*"[^>]*></script>', '', html_content)
        html_content = re.sub(r'<link[^>]*href="https?://[^"]*"[^>]*>', '', html_content)
        
        # Remove any remaining script tags that might cause issues
        html_content = re.sub(r'<script[^>]*>.*?</script>', '', html_content, flags=re.DOTALL)
        
        # The CSS contains problematic var() functions causing recursion
        # Skip the complex template and generate a clean PDF directly
        print("🔄 Generating clean PDF to avoid CSS recursion issues...", flush=True)
        
        # Create a clean, simple PDF-friendly HTML
        clean_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{company_name} - Company Profile</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 40px;
                    line-height: 1.6;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    border-bottom: 3px solid #007bff;
                    padding-bottom: 20px;
                    margin-bottom: 30px;
                }}
                h1 {{
                    color: #007bff;
                    font-size: 28px;
                    margin-bottom: 10px;
                }}
                h2 {{
                    color: #555;
                    font-size: 20px;
                    margin-top: 30px;
                    margin-bottom: 15px;
                    border-left: 4px solid #007bff;
                    padding-left: 15px;
                }}
                .section {{
                    margin: 25px 0;
                    padding: 15px;
                    border: 1px solid #eee;
                    border-radius: 5px;
                }}
                .metric {{
                    margin: 8px 0;
                    padding: 5px 0;
                }}
                .metric strong {{
                    color: #007bff;
                    display: inline-block;
                    width: 150px;
                }}
                .financial-grid {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 20px;
                    margin: 15px 0;
                }}
                .financial-item {{
                    padding: 10px;
                    background: #f8f9fa;
                    border-radius: 4px;
                }}
                .footer {{
                    margin-top: 40px;
                    padding-top: 20px;
                    border-top: 1px solid #ddd;
                    font-size: 12px;
                    color: #666;
                    text-align: center;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #f2f2f2;
                    font-weight: bold;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{company_name}</h1>
                <p>Company Profile Report</p>
            </div>

            <div class="section">
                <h2>Company Overview</h2>
                <div class="metric"><strong>Company Name:</strong> {company_name}</div>
                <div class="metric"><strong>Legal Form:</strong> {safe_get(company_overview.get('legal_form'), 'Not specified')}</div>
                <div class="metric"><strong>Founded:</strong> {safe_get(company_overview.get('companyFoundationyear'), 'Not specified')}</div>
                <div class="metric"><strong>Primary Sector:</strong> {safe_get(company_overview.get('primary_sector'), 'General sector')}</div>
                <div class="metric"><strong>Expertise:</strong> {safe_get(company_overview.get('companyExpertise'), 'To be determined')}</div>
            </div>"""
        
        # Add financial data if available
        if extracted_kpis:
            clean_html += """
            <div class="section">
                <h2>Financial Information</h2>
                <table>
                    <tr><th>Metric</th><th>Value</th></tr>"""
            
            for key, value in extracted_kpis.items():
                if value is not None:
                    clean_html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
            
            clean_html += "</table>"
        
        if computed_ratios:
            clean_html += """
                <h2>Financial Ratios</h2>
                <table>
                    <tr><th>Ratio</th><th>Value</th></tr>"""
            
            for key, value in computed_ratios.items():
                if value is not None:
                    clean_html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{value}</td></tr>"
            
            clean_html += "</table>"
        
        clean_html += """
            </div>

            <div class="section">
                <h2>Contact Information</h2>"""
        
        contact = basic_info.get('contact', {})
        clean_html += f"""
                <div class="metric"><strong>Address:</strong> {safe_get(contact.get('address'), 'Not available')}</div>
                <div class="metric"><strong>Phone:</strong> {safe_get(contact.get('phone'), 'Not available')}</div>
                <div class="metric"><strong>Email:</strong> {safe_get(contact.get('email'), 'Not available')}</div>
                <div class="metric"><strong>Website:</strong> {safe_get(contact.get('website'), 'Not available')}</div>
            </div>"""
        
        # Add sectors if available
        if basic_info.get('sectors'):
            clean_html += """
            <div class="section">
                <h2>Business Sectors</h2>"""
            sectors = basic_info.get('sectors')
            if sectors and isinstance(sectors, list):
                for sector in sectors:
                    if sector and isinstance(sector, dict):
                        title = safe_get(sector.get('title'), 'Sector')
                        description = safe_get(sector.get('description'), 'Description not available')
                        clean_html += f"<div class='metric'><strong>{title}:</strong> {description}</div>"
            clean_html += "</div>"
        
        # Add markets if available
        if basic_info.get('markets'):
            clean_html += """
            <div class="section">
                <h2>Target Markets</h2>"""
            markets = basic_info.get('markets')
            if markets and isinstance(markets, list):
                for market in markets:
                    if market and isinstance(market, dict):
                        title = safe_get(market.get('title'), 'Market')
                        description = safe_get(market.get('description'), 'Description not available')
                        clean_html += f"<div class='metric'><strong>{title}:</strong> {description}</div>"
            clean_html += "</div>"
        
        clean_html += f"""
            <div class="footer">
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>This is a simplified PDF version. For the complete interactive report with charts and detailed analysis, please view the HTML version.</p>
            </div>
        </body>
        </html>
        """
        
        # Generate PDF from clean HTML
        html_doc = HTML(string=clean_html, encoding='utf-8', base_url='')
        html_doc.write_pdf(pdf_buffer)
        
        # Get PDF content
        pdf_content = pdf_buffer.getvalue()
        pdf_buffer.close()
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y-%m-%d')
        filename = f"{company_name.replace(' ', '_')}_report_{timestamp}.pdf"
        
        # Return PDF response
        from flask import Response
        return Response(
            pdf_content,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Length': str(len(pdf_content))
            }
        )
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error generating PDF: {str(e)}", flush=True)
        print(f"❌ Full traceback: {error_details}", flush=True)
        return jsonify({'error': f'PDF generation failed: {str(e)}'}), 500

@app.route('/api/profiles/<profile_id>/send-email', methods=['POST'])
@jwt_required()
def send_profile_email(profile_id):
    """Manually send PDF report via email for a completed profile"""
    try:
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        # Check if profile is completed
        if profile.status != 'completed':
            return jsonify({'error': 'Profile must be completed before sending email'}), 400
        
        # Get the current user to check permissions
        user_id = get_jwt_identity()
        user = db.session.get(User, user_id)
        
        # Check if user is admin or the profile creator
        if user.role != 'admin' and profile.created_by != user_id:
            return jsonify({'error': 'Unauthorized to send email for this profile'}), 403
        
        # Send PDF report via email
        success = send_pdf_report_email(profile_id)
        
        if success:
            return jsonify({'message': 'PDF report sent successfully via email'}), 200
        else:
            return jsonify({'error': 'Failed to send PDF report via email'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/confirm-upload', methods=['POST'])
@jwt_required()
def confirm_upload_with_mismatch(profile_id):
    """
    Handle user confirmation when company names don't match.
    User can choose to proceed with the upload or cancel it.
    """
    try:
        # Check if profile exists
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        action = data.get('action')  # 'proceed' or 'cancel'
        if action not in ['proceed', 'cancel']:
            return jsonify({'error': 'Invalid action. Must be "proceed" or "cancel"'}), 400
        
        if action == 'cancel':
            return jsonify({
                'message': 'Upload cancelled by user due to company name mismatch',
                'action': 'cancelled'
            }), 200
        
        # If user chooses to proceed, we need the file data
        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'No files provided for confirmed upload'}), 400
        
        # Check file count
        if len(files) > 3:
            return jsonify({'error': 'Maximum 3 files allowed'}), 400
        
        # Proceed with upload (user has confirmed despite mismatch)
        uploaded_files = []
        
        for file in files:
            if file.filename == '':
                continue
                
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
            file.save(file_path)
            
            # Save document record
            document = LiasseDocument(
                profile_id=profile_id,
                file_name=filename,
                file_path=file_path,
                file_size=os.path.getsize(file_path)
            )
            
            db.session.add(document)
            uploaded_files.append({
                'id': document.id,
                'filename': filename,
                'size': document.file_size
            })
        
        db.session.commit()
        
        return jsonify({
            'message': 'Files uploaded successfully after user confirmation',
            'files': uploaded_files,
            'action': 'proceeded',
            'warning': 'Company names did not match but upload proceeded as confirmed by user'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/confirm-smart-upload', methods=['POST'])
@jwt_required()
def confirm_smart_upload_with_mismatch(profile_id):
    """
    Handle user confirmation when company names don't match during smart upload.
    User can choose to proceed with the smart upload or cancel it.
    """
    try:
        # Check if profile exists
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        action = data.get('action')  # 'proceed' or 'cancel'
        if action not in ['proceed', 'cancel']:
            return jsonify({'error': 'Invalid action. Must be "proceed" or "cancel"'}), 400
        
        if action == 'cancel':
            return jsonify({
                'message': 'Smart upload cancelled by user due to company name mismatch',
                'action': 'cancelled'
            }), 200
        
        # If user chooses to proceed, we need the file data
        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'No files provided for confirmed smart upload'}), 400
        
        # Check file count
        if len(files) > 3:
            return jsonify({'error': 'Maximum 3 files allowed'}), 400
        
        # Proceed with smart upload (user has confirmed despite mismatch)
        from services.profile_verification import identify_new_vs_existing_documents
        
        # Save files temporarily for analysis
        import tempfile
        temp_file_paths = []
        try:
            for i, file in enumerate(files):
                if file.filename == '':
                    continue
                    
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                    file.save(temp_file.name)
                    temp_file_paths.append(temp_file.name)
            
            # Extract company info from all documents
            from services.profile_verification import extract_company_info_from_first_page
            
            all_company_info = []
            api_key = app.config.get('ANTHROPIC_API_KEY')
            
            for i, file_path in enumerate(temp_file_paths):
                company_info = extract_company_info_from_first_page(file_path, api_key)
                if company_info:
                    all_company_info.append(company_info)
            
            # Now call the function with all required parameters
            document_analysis = identify_new_vs_existing_documents(
                db, CompanyProfile, temp_file_paths, profile.company_name, all_company_info
            )
            
            uploaded_files = []
            processed_count = 0
            
            # Handle existing documents - reuse saved data
            for match in document_analysis['existing_matches']:
                filename = os.path.basename(match['file_path'])
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
                
                # Copy the file to the new profile's directory
                import shutil
                shutil.copy2(match['file_path'], file_path)
                
                document = LiasseDocument(
                    profile_id=profile_id,
                    file_name=filename,
                    file_path=file_path,
                    file_size=os.path.getsize(file_path),
                    extracted_data=match['existing_data'].get('extracted_data'),
                    upload_status='reused',
                    ocr_status='completed'
                )
                
                db.session.add(document)
                uploaded_files.append({
                    'id': document.id,
                    'filename': filename,
                    'size': document.file_size,
                    'status': 'reused',
                    'message': 'Document data reused from existing profile'
                })
            
            # Handle new documents - process normally
            for new_doc in document_analysis['new_documents']:
                filename = os.path.basename(new_doc['file_path'])
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{profile_id}_{filename}")
                
                # Copy the file to the new profile's directory
                import shutil
                shutil.copy2(new_doc['file_path'], file_path)
                
                # Save document record for new document
                document = LiasseDocument(
                    profile_id=profile_id,
                    file_name=filename,
                    file_path=file_path,
                    file_size=os.path.getsize(file_path),
                    upload_status='uploaded',
                    ocr_status='pending'
                )
                
                db.session.add(document)
                uploaded_files.append({
                    'id': document.id,
                    'filename': filename,
                    'size': document.file_size,
                    'status': 'new',
                    'message': 'Document will be processed'
                })
                processed_count += 1
            
            db.session.commit()
            
            return jsonify({
                'message': 'Smart upload completed successfully after user confirmation',
                'uploaded_files': uploaded_files,
                'document_analysis': document_analysis,
                'new_documents_to_process': processed_count,
                'action': 'proceeded',
                'warning': 'Company names did not match but smart upload proceeded as confirmed by user'
            })
            
        finally:
            # Clean up temporary files
            for temp_path in temp_file_paths:
                try:
                    os.unlink(temp_path)
                except:
                    pass
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/profiles/<profile_id>/download-excel', methods=['GET'])
def download_excel_report(profile_id):
    """Download Excel version of the financial report with only financial numbers"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        profile = CompanyProfile.query.get_or_404(profile_id)
        
        if profile.status != 'completed':
            return jsonify({'error': 'Profile is not completed yet'}), 400
            
        profile_data = profile.profile_data or {}
        extracted_kpis = profile_data.get('extracted_kpis', {})
        computed_ratios = profile_data.get('computed_ratios', {})
        
        # Always use profile.company_name first (it's the source of truth and gets updated by re-scraping)
        # Check if profile_data has a placeholder, and if so, use profile.company_name instead
        company_name_from_data = profile_data.get('company_name')
        placeholder_values = ['company name placeholder', 'COMPANY NAME PLACEHOLDER', 
                             'Company Name Not Extracted', 'COMPANY NAME NOT EXTRACTED']
        
        if company_name_from_data and company_name_from_data.strip() in placeholder_values:
            # profile_data has placeholder, use profile.company_name (which is updated by re-scraping)
            company_name = profile.company_name
            print(f"📝 Excel Report: Using profile.company_name '{company_name}' instead of placeholder '{company_name_from_data}'", flush=True)
        else:
            # Use profile_data company_name if not a placeholder, otherwise fall back to profile.company_name
            company_name = company_name_from_data or profile.company_name
        
        # Clean the company name for display by removing legal forms
        from services.profile_verification import _normalize_company_name
        display_company_name = _normalize_company_name(company_name)
        
        # Final check: if display_company_name is still a placeholder after normalization, use profile.company_name
        normalized_placeholder_check = display_company_name.upper().strip()
        if normalized_placeholder_check in ['COMPANY NAME PLACEHOLDER', 'COMPANY NAME NOT EXTRACTED']:
            display_company_name = _normalize_company_name(profile.company_name)
            print(f"📝 Excel Report: Replaced placeholder '{normalized_placeholder_check}' with profile.company_name: '{display_company_name}'", flush=True)
        
        if not extracted_kpis and not computed_ratios:
            return jsonify({'error': 'No financial data available'}), 400
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Données Financières"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14, color="000000")
        data_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get available years
        available_years = []
        if computed_ratios and '_metadata' in computed_ratios:
            available_years = computed_ratios['_metadata'].get('available_years', ['N', 'N-1'])
        elif extracted_kpis and '_metadata' in extracted_kpis:
            available_years = extracted_kpis['_metadata'].get('available_years', ['N', 'N-1'])
        else:
            available_years = ['N', 'N-1']
        
        # Convert year labels to actual years if fiscal years are available
        actual_years = []
        if profile.fiscal_years:
            try:
                if '-' in str(profile.fiscal_years):
                    # Range format like "2022-2023"
                    years = str(profile.fiscal_years).split('-')
                    current_year = int(years[1]) if len(years) > 1 else int(years[0])
                    previous_year = int(years[0])
                    actual_years = [str(previous_year), str(current_year)]
                else:
                    # Single year format
                    current_year = int(profile.fiscal_years)
                    previous_year = current_year - 1
                    actual_years = [str(previous_year), str(current_year)]
            except:
                actual_years = available_years
        else:
            actual_years = available_years
        
        # Headers - Dynamic based on available years
        headers = ['Indicateur Financier', 'Type']
        # Add year columns dynamically
        for year in actual_years:
            headers.append(f'{year} (MAD)')
        
        # Title row - merge cells based on number of columns
        title_range = f'A1:{get_column_letter(len(headers))}1'
        ws.merge_cells(title_range)
        ws['A1'] = f"Données Financières - {display_company_name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Define financial KPIs to include
        financial_kpis = [
            {
                'name': "Chiffre d'affaires",
                'key': "Chiffre d'affaires",
                'type': 'Extrait'
            },
            {
                'name': 'Résultat Net',
                'key': 'Résultat Net',
                'type': 'Extrait'
            },
            {
                'name': 'Résultat d\'exploitation',
                'key': 'Résultat d\'exploitation',
                'type': 'Extrait'
            },
            {
                'name': 'Capitaux propres',
                'key': 'Capitaux propres',
                'type': 'Extrait'
            },
            {
                'name': 'Dettes de financement',
                'key': 'Dettes de financement',
                'type': 'Extrait'
            },
            {
                'name': 'Trésorerie-Actif',
                'key': 'Trésorerie-Actif',
                'type': 'Extrait'
            },
            {
                'name': 'Trésorerie-passif',
                'key': 'Trésorerie-passif',
                'type': 'Extrait'
            },
            {
                'name': 'Actif circulant',
                'key': 'Actif circulant',
                'type': 'Extrait'
            },
            {
                'name': 'Passif circulant',
                'key': 'Passif circulant',
                'type': 'Extrait'
            },
            {
                'name': 'EBITDA',
                'key': 'ebitda',
                'type': 'Calculé'
            },
            {
                'name': 'Dette Nette',
                'key': 'dette_nette',
                'type': 'Calculé'
            },
            {
                'name': 'Trésorerie Nette',
                'key': 'tresorerie_nette',
                'type': 'Calculé'
            },
            {
                'name': 'BFR',
                'key': 'bfr',
                'type': 'Calculé'
            },
            {
                'name': 'Marge EBITDA (%)',
                'key': 'marge_ebitda',
                'type': 'Calculé'
            },
            {
                'name': 'Marge d\'exploitation (%)',
                'key': 'marge_exploitation',
                'type': 'Calculé'
            },
            {
                'name': 'Marge Nette (%)',
                'key': 'marge_nette',
                'type': 'Calculé'
            },
            {
                'name': 'ROE (%)',
                'key': 'roe',
                'type': 'Calculé'
            },
            {
                'name': 'ROCE (%)',
                'key': 'roce',
                'type': 'Calculé'
            },
            {
                'name': 'Gearing (%)',
                'key': 'gearing',
                'type': 'Calculé'
            },
            {
                'name': 'Capacité de remboursement',
                'key': 'capacite_remboursements',
                'type': 'Calculé'
            }
        ]
        
        # Helper function to get KPI value
        def get_kpi_value(kpi_key, year_label):
            # Try computed ratios first (flat structure)
            flat_key = f"{kpi_key}_{year_label.lower()}"
            if flat_key in computed_ratios:
                return computed_ratios[flat_key]
            
            # Try extracted KPIs (nested structure)
            if kpi_key in extracted_kpis and isinstance(extracted_kpis[kpi_key], dict):
                return extracted_kpis[kpi_key].get(year_label)
            
            return None
        
        # Populate data - One row per KPI with years as columns
        row = 4
        for kpi in financial_kpis:
            # Write KPI name and type
            ws.cell(row=row, column=1, value=kpi['name']).font = data_font
            ws.cell(row=row, column=2, value=kpi['type']).font = data_font
            
            # Write values for each year in separate columns
            for i, year_label in enumerate(available_years):
                actual_year = actual_years[i] if i < len(actual_years) else year_label
                value = get_kpi_value(kpi['key'], year_label)
                
                # Format value
                if value is not None:
                    if isinstance(value, (int, float)):
                        if kpi['type'] == 'Calculé' and '%' in kpi['name']:
                            formatted_value = f"{value:.2f}%"
                        else:
                            formatted_value = f"{value:,.0f}"
                    else:
                        formatted_value = str(value)
                else:
                    formatted_value = "N/A"
                
                # Write value in the corresponding year column (column 3 + i)
                ws.cell(row=row, column=3 + i, value=formatted_value).font = data_font
            
            # Apply borders to all columns
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return Excel file
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"{display_company_name}_donnees_financieres.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Error generating Excel report: {e}", flush=True)
        return jsonify({'error': f'Error generating Excel report: {str(e)}'}), 500

# ============================================================================
# BLUEPRINT REGISTRATION
# ============================================================================
# Import and register benchmark routes
try:
    from benchmark.routes import benchmark_bp, init_benchmark_routes
    # Initialize routes with database instances
    init_benchmark_routes(db, CompanyProfile)
    app.register_blueprint(benchmark_bp)
    print("✅ Benchmark routes registered successfully")
            
except ImportError as e:
    print(f"⚠️ Could not import benchmark routes: {e}")
except Exception as e:
    print(f"❌ Error registering benchmark routes: {e}")

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================
if __name__ == '__main__':
    # Wait for database to be ready
    wait_for_db()
    
    with app.app_context():
        db.create_all()
        print("Database tables created successfully")
    
    print("Starting Flask application...")
    app.run(host='0.0.0.0', port=5000, debug=True)
